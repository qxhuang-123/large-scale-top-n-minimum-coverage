from __future__ import annotations

import json
import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd

try:
    import matplotlib.pyplot as plt
except Exception:
    plt = None
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, Side

ROOT = Path(r"C:\Users\24qxh\Documents\Codex\2026-07-05\import-json-import-sys-import-time-4")
OLD_ROOT = Path(r"C:\Users\24qxh\Documents\Codex\2026-07-04\new-chat")
CPP_DIR = ROOT / "work" / "tpcar_cpp39_history"
CACHE = OLD_ROOT / "work" / "yelp_cache"
OUT_DIR = ROOT / "outputs"
OUT_JSON = OUT_DIR / "yelp_exact_N10_D_percent_history_results.json"
OUT_XLSX = OUT_DIR / "yelp_exact_N10_D_percent_history_results.xlsx"
OUT_FIG = OUT_DIR / "yelp_exact_N10_iteration_curves.png"

sys.path.insert(0, str(CPP_DIR))
import tpcar_core_fast as tpcar_core  # noqa: E402

N = 10
D_PCTS = [0, 20, 40, 60, 80, 100]
PROGRESS_EVERY = 200


def build_cpp_arrays(scores, z):
    user_indptr = z["user_indptr"].astype(np.int32, copy=False)
    user_items = z["user_indices"].astype(np.int32, copy=False)
    item_indptr = z["item_indptr"].astype(np.int32, copy=False)
    item_users = z["item_users"].astype(np.int32, copy=False)

    num_users, num_items = scores.shape
    num_edges = len(user_items)

    edge_users = np.empty(num_edges, dtype=np.int32)
    user_scores = np.empty(num_edges, dtype=np.float64)

    for u in range(num_users):
        start, end = int(user_indptr[u]), int(user_indptr[u + 1])
        if end > start:
            items = user_items[start:end]
            edge_users[start:end] = u
            user_scores[start:end] = scores[u, items]

    item_edges = np.empty(num_edges, dtype=np.int32)
    cursor = item_indptr[:-1].copy()

    for edge_id, item in enumerate(user_items):
        pos = cursor[item]
        item_edges[pos] = edge_id
        cursor[item] += 1

    item_scores = user_scores[item_edges]

    return (
        user_indptr,
        user_items,
        user_scores,
        edge_users,
        item_indptr,
        item_users,
        item_edges,
        item_scores,
    )


def d_target_from_pct(d_base: int, pct: int) -> int:
    return int(math.ceil(d_base * pct / 100.0))


def choose_first_history_at_or_after(history: list[dict], target_d: int) -> dict:
    for row in history:
        if int(row["diversity"]) >= target_d:
            return row
    return history[-1]


def make_percent_rows(result: dict, d_base: int, total_elapsed: float) -> list[dict]:
    history = list(result["history"])
    total_recs = int(result["total_recs"])
    rows = []
    for pct in D_PCTS:
        target_d = d_target_from_pct(d_base, pct)
        h = choose_first_history_at_or_after(history, target_d)
        objective = float(h["objective"])
        rows.append(
            {
                "D比例": f"D-{pct}%",
                "D_target": target_d,
                "D_target_base": "items_with_candidates",
                "迭代次数": int(h["iteration"]),
                "总体多样性": int(h["diversity"]),
                "目标函数值": round(objective, 6),
                "准确多样性": round(objective / total_recs, 6) if total_recs else 0.0,
                "时间s": round(float(h["time_sec"]), 3),
                "target_feasible": bool(int(h["diversity"]) >= target_d),
            }
        )
    rows[-1]["时间s"] = round(total_elapsed, 3)
    return rows


def make_history_rows(result: dict, total_recs: int) -> list[dict]:
    rows = []
    for h in result["history"]:
        objective = float(h["objective"])
        rows.append(
            {
                "迭代次数": int(h["iteration"]),
                "总体多样性": int(h["diversity"]),
                "目标函数值": round(objective, 6),
                "时间s": round(float(h["time_sec"]), 3),
                "准确多样性": round(objective / total_recs, 6) if total_recs else 0.0,
            }
        )
    return rows


def write_excel(meta: dict, percent_rows: list[dict], history_rows: list[dict], summary_rows: list[dict]):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    percent_df = pd.DataFrame(percent_rows)
    history_df = pd.DataFrame(history_rows)
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        history_df[["迭代次数", "总体多样性", "目标函数值", "时间s"]].to_excel(
            writer, sheet_name="迭代明细", index=False
        )
        percent_df.to_excel(writer, sheet_name="D比例结果", index=False)
        summary_df.to_excel(writer, sheet_name="总结果", index=False)
        pd.DataFrame([meta]).to_excel(writer, sheet_name="meta", index=False)

    wb = load_workbook(OUT_XLSX)
    thin = Side(style="thin", color="000000")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 22)

    ws = wb["迭代明细"]
    if ws.max_row >= 2:
        chart1 = LineChart()
        chart1.title = "总体多样性迭代曲线"
        chart1.y_axis.title = "总体多样性"
        chart1.x_axis.title = "迭代次数"
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 7
        chart1.width = 14
        ws.add_chart(chart1, "F2")

        chart2 = LineChart()
        chart2.title = "目标函数值迭代曲线"
        chart2.y_axis.title = "目标函数值"
        chart2.x_axis.title = "迭代次数"
        data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 7
        chart2.width = 14
        ws.add_chart(chart2, "F18")

    wb.save(OUT_XLSX)


def write_figure(history_rows: list[dict]):
    if plt is None:
        print("matplotlib not installed; PNG figure skipped. Excel charts are still written.", flush=True)
        return
    if not history_rows:
        return
    df = pd.DataFrame(history_rows).sort_values("迭代次数")
    fig, axes = plt.subplots(1, 2, figsize=(12, 4), dpi=160)
    axes[0].plot(df["迭代次数"], df["总体多样性"], linewidth=1.8)
    axes[0].set_title("Yelp N=10 总体多样性")
    axes[0].set_xlabel("迭代次数")
    axes[0].set_ylabel("总体多样性 D")
    axes[0].grid(True, alpha=0.25)
    axes[1].plot(df["迭代次数"], df["目标函数值"], linewidth=1.8, color="#b45f06")
    axes[1].set_title("Yelp N=10 目标函数值")
    axes[1].set_xlabel("迭代次数")
    axes[1].set_ylabel("目标函数值")
    axes[1].grid(True, alpha=0.25)
    fig.tight_layout()
    fig.savefig(OUT_FIG)
    plt.close(fig)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("loading scores/candidates", flush=True)
    scores = np.load(CACHE / "yelp_unknown_scores_float32.npy")
    z = np.load(CACHE / "yelp_cand_frac_0p4_seed20260704.npz")

    num_users, num_items = scores.shape
    candidate_edges = int(len(z["user_indices"]))
    items_with_candidates = int((np.diff(z["item_indptr"]) > 0).sum())
    max_d_target = items_with_candidates
    print(
        "shape",
        (num_users, num_items),
        "candidate_edges",
        candidate_edges,
        "items_with_candidates",
        items_with_candidates,
        flush=True,
    )

    print("building arrays for C++", flush=True)
    t0 = time.time()
    arrays = build_cpp_arrays(scores, z)
    array_build_sec = time.time() - t0
    print("array_build_sec", round(array_build_sec, 3), flush=True)

    meta = {
        "dataset": "Yelp",
        "N": N,
        "num_users": int(num_users),
        "num_items": int(num_items),
        "candidate_edges": candidate_edges,
        "items_with_candidates": items_with_candidates,
        "D_percentages": D_PCTS,
        "D_target_base": "items_with_candidates",
        "engine": "cpp_pybind11_exact_with_history",
        "cpp_module_dir": str(CPP_DIR),
        "cache_dir": str(CACHE),
        "array_build_sec": round(array_build_sec, 3),
    }

    print(f"--- EXACT N={N}, D={max_d_target} ---", flush=True)
    t0 = time.time()
    raw = dict(
        tpcar_core.run_exact_csr(
            int(num_users),
            int(num_items),
            *arrays,
            int(N),
            int(max_d_target),
            int(PROGRESS_EVERY),
        )
    )
    total_elapsed = time.time() - t0

    history_rows = make_history_rows(raw, int(raw["total_recs"]))
    percent_rows = make_percent_rows(raw, items_with_candidates, total_elapsed)
    summary_rows = [
        {
            "N": N,
            "总时间s": round(total_elapsed, 3),
            "目标函数值": round(float(raw["tpcar_obj"]), 6),
            "准确多样性": round(float(raw["tpcar_pred"]), 6),
            "总体多样性": int(raw["tpcar_D"]),
            "Naive目标函数值": round(float(raw["naive_obj"]), 6),
            "Naive准确多样性": round(float(raw["naive_pred"]), 6),
            "Naive总体多样性": int(raw["naive_D"]),
            "迭代次数": int(raw["augmentations"]),
            "交换次数": int(raw["swaps"]),
            "target_feasible": bool(raw["target_feasible"]),
            "max_reached": int(raw["max_reached"]),
            "total_recs": int(raw["total_recs"]),
        }
    ]

    payload = {
        "meta": meta,
        "D_percent_results": percent_rows,
        "summary": summary_rows,
        "history": history_rows,
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(meta, percent_rows, history_rows, summary_rows)
    write_figure(history_rows)

    print("D percent results:", flush=True)
    for row in percent_rows:
        print(row, flush=True)
    print("summary:", summary_rows[0], flush=True)
    print(OUT_JSON, flush=True)
    print(OUT_XLSX, flush=True)
    if OUT_FIG.exists():
        print(OUT_FIG, flush=True)
    else:
        print("PNG skipped; Excel charts are included in the workbook.", flush=True)


if __name__ == "__main__":
    main()

