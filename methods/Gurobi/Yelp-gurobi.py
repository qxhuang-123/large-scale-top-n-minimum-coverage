from __future__ import annotations

import argparse
import json
import math
import os
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from scipy import sparse

try:
    import gurobipy as gp
    from gurobipy import GRB
except ImportError as exc:
    raise SystemExit(
        "Cannot import gurobipy. Install Gurobi's Python package in the Python "
        "environment used to run this script, then rerun.\n"
        "Example: python -m pip install gurobipy"
    ) from exc


# ------------------------- user settings -------------------------

INPUT_XLSX = Path(r"E:\Users\24qxh\Desktop\Yelp_R_ui_UNKNOWN_ONLY_known_ratings_0_display1.xlsx")
SHEET_NAME = None  # None means use the first sheet.

DATASET_DIR = Path(r"E:\Users\24qxh\Desktop\LETTER-main\LETTER-main\dataset\Yelp")
MASK_KNOWN_RATINGS = True

N = 10
ALPHA = 0.40
D_PERCENTAGES = [0, 20, 40, 60, 80, 100]

USE_LP_RELAXATION = False
TIME_LIMIT_SECONDS = None  # Example: 3600. None means no time limit.
THREADS = 0  # 0 lets Gurobi decide.
MIP_GAP = 0.0

OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT_XLSX = OUTPUT_DIR / "op_gurobi_binary_mip_N10_D_percent_history_results.xlsx"
OUTPUT_PNG = OUTPUT_DIR / "op_gurobi_binary_mip_N10_D_percent_iteration_plot.png"


# ------------------------- data loading -------------------------

def load_prediction_matrix(xlsx_path: Path, sheet_name: str | None) -> tuple[np.ndarray, list[str], list[str]]:
    print(f"Loading Excel matrix: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    item_names = [str(v) for v in headers[1:]]

    num_users = ws.max_row - 1
    num_items = ws.max_column - 1
    ratings = np.empty((num_users, num_items), dtype=np.float32)
    user_ids: list[str] = []

    t0 = time.time()
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        user_ids.append(str(row[0]))
        values = row[1:]
        ratings[r_idx, :] = [
            np.nan if v is None or v == "" else float(v)
            for v in values
        ]
        if (r_idx + 1) % 500 == 0:
            print(f"  read {r_idx + 1:,}/{num_users:,} users")

    print(
        f"Loaded R shape={ratings.shape}, "
        f"finite={np.isfinite(ratings).sum():,}, time={time.time() - t0:.1f}s"
    )
    return ratings, user_ids, item_names


def apply_known_rating_mask(ratings: np.ndarray, dataset_dir: Path) -> int:
    if not MASK_KNOWN_RATINGS:
        return 0
    if not dataset_dir.exists():
        print(f"Known-rating mask skipped: dataset dir not found: {dataset_dir}")
        return 0

    num_users, num_items = ratings.shape
    masked = 0
    for split in ["train", "val", "test"]:
        path = dataset_dir / f"{split}.json"
        if not path.exists():
            print(f"Known-rating mask skipped for missing file: {path}")
            continue
        print(f"Applying known-rating mask: {path}")
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        for uid_str, items in data.items():
            uid = int(uid_str)
            if uid < 0 or uid >= num_users:
                continue
            for pair in items:
                iid = int(pair[1])
                if 0 <= iid < num_items and np.isfinite(ratings[uid, iid]):
                    ratings[uid, iid] = np.nan
                    masked += 1
    print(f"Masked known ratings: {masked:,}")
    return masked


def build_candidate_edges(ratings: np.ndarray, alpha: float, n: int):
    num_users, num_items = ratings.shape
    edge_u_parts: list[np.ndarray] = []
    edge_i_parts: list[np.ndarray] = []
    edge_score_parts: list[np.ndarray] = []
    user_quota = np.zeros(num_users, dtype=np.int32)

    t0 = time.time()
    for u in range(num_users):
        row = ratings[u]
        valid = np.flatnonzero(np.isfinite(row))
        if valid.size == 0:
            continue

        k = max(n, int(math.ceil(alpha * valid.size)))
        k = min(k, valid.size)

        scores = row[valid]
        if k < valid.size:
            local = np.argpartition(scores, -k)[-k:]
            cand = valid[local]
        else:
            cand = valid

        cand_scores = row[cand].astype(np.float64, copy=False)
        order = np.argsort(-cand_scores, kind="mergesort")
        cand = cand[order].astype(np.int32, copy=False)
        cand_scores = cand_scores[order]

        edge_u_parts.append(np.full(cand.size, u, dtype=np.int32))
        edge_i_parts.append(cand)
        edge_score_parts.append(cand_scores)
        user_quota[u] = min(n, cand.size)

        if (u + 1) % 500 == 0:
            print(f"  candidates for {u + 1:,}/{num_users:,} users")

    edge_u = np.concatenate(edge_u_parts)
    edge_i = np.concatenate(edge_i_parts)
    edge_score = np.concatenate(edge_score_parts).astype(np.float64, copy=False)

    item_has_candidate = np.bincount(edge_i, minlength=num_items) > 0
    print(
        f"Candidate edges={edge_u.size:,}, items_with_candidates={item_has_candidate.sum():,}/"
        f"{num_items:,}, total_slots={user_quota.sum():,}, time={time.time() - t0:.1f}s"
    )
    return edge_u, edge_i, edge_score, user_quota, item_has_candidate


def compute_naive(edge_u, edge_i, edge_score, user_quota, num_users, num_items):
    selected = np.zeros(edge_u.size, dtype=bool)
    start = 0
    for u in range(num_users):
        end = start
        while end < edge_u.size and edge_u[end] == u:
            end += 1
        k = int(user_quota[u])
        if k > 0:
            selected[start:start + k] = True
        start = end

    item_counts = np.bincount(edge_i[selected], minlength=num_items)
    obj = float(edge_score[selected].sum())
    total_slots = int(user_quota.sum())
    return {
        "objective": obj,
        "accuracy": obj / total_slots if total_slots else 0.0,
        "coverage": int((item_counts > 0).sum()),
    }


# ------------------------- Gurobi model -------------------------

def build_gurobi_model(edge_u, edge_i, edge_score, user_quota, num_users, num_items):
    num_edges = edge_u.size
    num_vars = num_edges + num_items
    vtype = GRB.CONTINUOUS if USE_LP_RELAXATION else GRB.BINARY

    print(f"Building Gurobi model: vars={num_vars:,}, edge_vars={num_edges:,}, z_vars={num_items:,}")
    model = gp.Model("OP_TopN_Coverage_Gurobi")
    y = model.addMVar(num_vars, lb=0.0, ub=1.0, vtype=vtype, name="y")

    obj = np.zeros(num_vars, dtype=np.float64)
    obj[:num_edges] = edge_score
    model.setObjective(obj @ y, GRB.MAXIMIZE)

    row_user = edge_u
    col_user = np.arange(num_edges, dtype=np.int64)
    data_user = np.ones(num_edges, dtype=np.float64)
    a_user = sparse.csr_matrix(
        (data_user, (row_user, col_user)),
        shape=(num_users, num_vars),
    )

    row_link = np.concatenate([edge_i, np.arange(num_items, dtype=np.int32)])
    col_link = np.concatenate([
        np.arange(num_edges, dtype=np.int64),
        num_edges + np.arange(num_items, dtype=np.int64),
    ])
    data_link = np.concatenate([
        -np.ones(num_edges, dtype=np.float64),
        np.ones(num_items, dtype=np.float64),
    ])
    a_link = sparse.csr_matrix(
        (data_link, (row_link, col_link)),
        shape=(num_items, num_vars),
    )

    a_d = sparse.csr_matrix(
        (
            np.ones(num_items, dtype=np.float64),
            (np.zeros(num_items, dtype=np.int32), num_edges + np.arange(num_items, dtype=np.int64)),
        ),
        shape=(1, num_vars),
    )

    a = sparse.vstack([a_user, a_link, a_d], format="csr")
    sense = np.array(["="] * num_users + ["<"] * num_items + [">"], dtype="U1")
    rhs = np.concatenate([
        user_quota.astype(np.float64),
        np.zeros(num_items, dtype=np.float64),
        np.array([0.0], dtype=np.float64),
    ])

    print(f"Adding sparse constraints: rows={a.shape[0]:,}, nnz={a.nnz:,}")
    model.addMConstr(a, y, sense, rhs, name="constraints")

    if USE_LP_RELAXATION:
        model.Params.Method = 1  # dual simplex, returns a basic integer solution for TU models.
    else:
        model.Params.MIPGap = MIP_GAP
    if TIME_LIMIT_SECONDS is not None:
        model.Params.TimeLimit = TIME_LIMIT_SECONDS
    if THREADS is not None:
        model.Params.Threads = THREADS

    model.Params.OutputFlag = 1
    model.update()
    return model, y, rhs


def solve_for_d(model, y, rhs, d_constr, d_target: int, num_edges: int, edge_i, num_items: int):
    constr = d_constr
    constr.RHS = float(d_target)
    model.update()

    t0 = time.time()
    model.optimize()
    solve_time = time.time() - t0

    status = model.Status
    if status not in (GRB.OPTIMAL, GRB.TIME_LIMIT, GRB.SUBOPTIMAL):
        return {
            "status": status,
            "solve_time": solve_time,
            "objective": None,
            "accuracy": None,
            "coverage": None,
            "fractional_x": None,
            "gap": None,
        }

    sol = np.asarray(y.X)
    x_sol = sol[:num_edges]
    selected = x_sol > 0.5
    item_counts = np.bincount(edge_i[selected], minlength=num_items)
    objective = float(model.ObjVal)
    fractional_x = int(((x_sol > 1e-6) & (x_sol < 1 - 1e-6)).sum())
    gap = None
    if not USE_LP_RELAXATION and hasattr(model, "MIPGap"):
        gap = float(model.MIPGap)

    return {
        "status": status,
        "solve_time": solve_time,
        "objective": objective,
        "accuracy": None,
        "coverage": int((item_counts > 0).sum()),
        "fractional_x": fractional_x,
        "gap": gap,
    }


# ------------------------- output -------------------------

def write_results(workbook_path: Path, iteration_rows, d_rows, summary_row, meta_row):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "迭代明细"
    ws.append(["迭代次数", "总体多样性", "目标函数值", "时间s"])
    for row in iteration_rows:
        ws.append(row)

    chart = LineChart()
    chart.title = "迭代次数-总体多样性"
    chart.y_axis.title = "总体多样性"
    chart.x_axis.title = "迭代次数"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")

    ws2 = wb.create_sheet("D比例结果")
    ws2.append(["D比例", "D_target", "迭代次数", "总体多样性", "目标函数值", "准确多样性", "时间s", "target_feasible"])
    for row in d_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("总结果")
    ws3.append([
        "N", "总时间s", "目标函数值", "准确多样性", "总体多样性",
        "Naive目标函数值", "Naive准确多样性", "Naive总体多样性",
        "迭代次数", "交换次数", "target_feasible", "max_reached",
    ])
    ws3.append(summary_row)

    ws4 = wb.create_sheet("meta")
    ws4.append(list(meta_row.keys()))
    ws4.append(list(meta_row.values()))

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for col_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 45)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    print(f"Saved Excel: {workbook_path}")


def try_write_png(iteration_rows, png_path: Path):
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        print("matplotlib not installed; PNG plot skipped. The Excel file contains an embedded chart.")
        return

    x = [r[0] for r in iteration_rows]
    cov = [r[1] for r in iteration_rows]
    obj = [r[2] for r in iteration_rows]

    fig, ax1 = plt.subplots(figsize=(8, 4.8), dpi=160)
    ax1.plot(x, cov, marker="o", color="#1f77b4", label="总体多样性")
    ax1.set_xlabel("迭代次数")
    ax1.set_ylabel("总体多样性", color="#1f77b4")
    ax1.tick_params(axis="y", labelcolor="#1f77b4")

    ax2 = ax1.twinx()
    ax2.plot(x, obj, marker="s", color="#d62728", label="目标函数值")
    ax2.set_ylabel("目标函数值", color="#d62728")
    ax2.tick_params(axis="y", labelcolor="#d62728")

    fig.tight_layout()
    fig.savefig(png_path)
    plt.close(fig)
    print(f"Saved PNG: {png_path}")


def check_environment_only():
    print("Python:", sys.executable)
    print("gurobipy:", gp.gurobi.version())
    print("GUROBI_HOME:", os.environ.get("GUROBI_HOME", ""))
    print("INPUT_XLSX exists:", INPUT_XLSX.exists(), INPUT_XLSX)
    if INPUT_XLSX.exists():
        wb = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]
        print("sheet:", ws.title)
        print("rows:", ws.max_row, "cols:", ws.max_column, "items:", ws.max_column - 1)
    try:
        env = gp.Env(empty=True)
        env.setParam("OutputFlag", 0)
        env.start()
        m = gp.Model(env=env)
        x = m.addVar(lb=0, ub=1, obj=1)
        m.ModelSense = GRB.MAXIMIZE
        m.optimize()
        print("Gurobi license/model test: OK")
    except Exception as exc:
        print("Gurobi license/model test: FAILED")
        print(type(exc).__name__ + ":", exc)
    print("Check-only finished. Full optimization was not started.")


def main():
    parser = argparse.ArgumentParser(description='Run OP Gurobi binary MIP Top-N coverage solver.')
    parser.add_argument('--check-only', action='store_true', help='check Python/Gurobi/input file only, do not optimize')
    args = parser.parse_args()
    if args.check_only:
        check_environment_only()
        return

    start_all = time.time()
    ratings, user_ids, item_names = load_prediction_matrix(INPUT_XLSX, SHEET_NAME)
    num_users, num_items = ratings.shape
    masked_known = apply_known_rating_mask(ratings, DATASET_DIR)

    edge_u, edge_i, edge_score, user_quota, item_has_candidate = build_candidate_edges(ratings, ALPHA, N)
    total_slots = int(user_quota.sum())
    naive = compute_naive(edge_u, edge_i, edge_score, user_quota, num_users, num_items)

    model, y, rhs = build_gurobi_model(edge_u, edge_i, edge_score, user_quota, num_users, num_items)
    num_edges = edge_u.size
    d_constr = model.getConstrs()[num_users + num_items]
    max_candidate_items = int(item_has_candidate.sum())

    iteration_rows = []
    d_rows = []
    last_good = None

    for iteration, pct in enumerate(D_PERCENTAGES):
        raw_target = int(math.ceil(max_candidate_items * pct / 100.0))
        d_target = min(raw_target, max_candidate_items)
        target_feasible_by_candidates = raw_target <= max_candidate_items

        print("\n" + "=" * 72)
        print(f"Solving D-{pct}%: raw_target={raw_target}, model_target={d_target}")
        result = solve_for_d(
            model, y, rhs, d_constr, d_target,
            num_edges, edge_i, num_items,
        )
        elapsed_total = time.time() - start_all

        if result["objective"] is None:
            print(f"  status={result['status']}; no feasible solution for target {d_target}")
            d_rows.append([f"D-{pct}%", raw_target, iteration, None, None, None, round(elapsed_total, 3), False])
            continue

        objective = float(result["objective"])
        accuracy = objective / total_slots if total_slots else 0.0
        coverage = int(result["coverage"])
        target_feasible = bool(target_feasible_by_candidates and coverage >= raw_target)

        row_iter = [iteration, coverage, round(objective, 6), round(elapsed_total, 3)]
        iteration_rows.append(row_iter)
        d_rows.append([
            f"D-{pct}%", raw_target, iteration, coverage,
            round(objective, 6), round(accuracy, 6),
            round(elapsed_total, 3), target_feasible,
        ])

        last_good = {
            "objective": objective,
            "accuracy": accuracy,
            "coverage": coverage,
            "iteration": iteration,
            "target_feasible": target_feasible,
        }

        print(
            f"  coverage={coverage:,}, objective={objective:.6f}, "
            f"accuracy={accuracy:.6f}, elapsed_total={elapsed_total:.1f}s, "
            f"fractional_x={result['fractional_x']}"
        )

    if last_good is None:
        raise RuntimeError("No feasible Gurobi solution was produced.")

    total_time = time.time() - start_all
    summary_row = [
        N,
        round(total_time, 3),
        round(last_good["objective"], 6),
        round(last_good["accuracy"], 6),
        last_good["coverage"],
        round(naive["objective"], 6),
        round(naive["accuracy"], 6),
        naive["coverage"],
        last_good["iteration"],
        None,
        last_good["target_feasible"],
        last_good["coverage"],
    ]

    meta_row = {
        "dataset": "OP",
        "N": N,
        "alpha": ALPHA,
        "num_users": num_users,
        "num_items": num_items,
        "candidate_edges": int(edge_u.size),
        "items_with_candidates": max_candidate_items,
        "total_slots": total_slots,
        "D_percentages": str(D_PERCENTAGES),
        "engine": "gurobipy_binary_mip",
        "input_xlsx": str(INPUT_XLSX),
        "mask_known_ratings": MASK_KNOWN_RATINGS,
        "masked_known_ratings": masked_known,
        "use_lp_relaxation": USE_LP_RELAXATION,
        "output_xlsx": str(OUTPUT_XLSX),
        "total_time_sec": round(total_time, 3),
    }

    write_results(OUTPUT_XLSX, iteration_rows, d_rows, summary_row, meta_row)
    try_write_png(iteration_rows, OUTPUT_PNG)
    print("\nDone.")


if __name__ == "__main__":
    main()