from __future__ import annotations

import argparse
import json
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

try:
    from numba import njit
except Exception:
    njit = None


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_N = 10
D_PERCENTAGES = [0, 20, 40, 60, 80, 100]
SEED = 20260704
DEFAULT_HISTORY_INTERVAL = 50
DEFAULT_MAX_ROUNDS = 1000

DATASETS = {
    "OP": {
        "scores": Path(r"C:\Users\24qxh\Documents\Codex\2026-07-05\import-json-import-sys-import-time-4\work\op_exact_cache\op_full_scores_float32.npy"),
        "candidates": Path(r"C:\Users\24qxh\Documents\Codex\2026-07-05\import-json-import-sys-import-time-4\work\op_exact_cache\op_full_cand_frac_0p4_seed20260704.npz"),
        "data_path": Path(r"E:\Users\24qxh\Desktop\op\IUIC1_Rui_pred_user_item_FULL_display1.xlsx"),
        "candidate_fraction": 0.4,
    },
    "Yelp": {
        "scores": Path(r"C:\Users\24qxh\Documents\Codex\2026-07-04\new-chat\work\yelp_cache\yelp_unknown_scores_float32.npy"),
        "candidates": Path(r"C:\Users\24qxh\Documents\Codex\2026-07-04\new-chat\work\yelp_cache\yelp_cand_frac_0p4_seed20260704.npz"),
        "data_path": Path(r"E:\Users\24qxh\Desktop\Yelp_R_ui_UNKNOWN_ONLY_known_ratings_0_display1.xlsx"),
        "candidate_fraction": 0.4,
    },
    "VG": {
        "scores": SCRIPT_DIR / "VG贪心算法_cache" / "VG_scores_float32.npy",
        "candidates": SCRIPT_DIR / "VG贪心算法_cache" / "VG_cand_pos_top40_seed20260704_v2.npz",
        "data_path": Path(r"E:\Users\24qxh\Desktop\VG\VG_R_ui_FULL_display1.xlsx"),
        "candidate_fraction": 0.4,
    },
    "TG": {
        "scores": SCRIPT_DIR / "TG贪心算法_cache" / "TG_scores_float32.npy",
        "candidates": SCRIPT_DIR / "TG贪心算法_cache" / "TG_cand_pos_top40_seed20260704_v2.npz",
        "data_path": Path(r"C:\Users\24qxh\Documents\Codex\2026-07-11\new-chat-2\outputs"),
        "candidate_fraction": 0.4,
    },
    "SO": {
        "scores": SCRIPT_DIR / "SO贪心算法_cache" / "SO_scores_float32.npy",
        "candidates": SCRIPT_DIR / "SO贪心算法_cache" / "SO_cand_pos_top40_seed20260704_v2.npz",
        "data_path": Path(r"E:\Users\24qxh\Desktop\SO"),
        "candidate_fraction": 0.4,
    },
}


if njit is not None:

    @njit(cache=True)
    def _initial_topn_numba(num_users, num_items, user_indptr, user_items, user_scores, n):
        selected_items = np.full((num_users, n), -1, dtype=np.int32)
        selected_scores = np.zeros((num_users, n), dtype=np.float32)
        counts = np.zeros(num_items, dtype=np.int32)
        total_score = 0.0
        total_recs = 0

        for u in range(num_users):
            s = user_indptr[u]
            e = user_indptr[u + 1]
            length = e - s
            if length <= 0:
                continue
            k = n if n < length else length
            order = np.argsort(user_scores[s:e])[::-1]
            for pos in range(k):
                edge = s + order[pos]
                item = user_items[edge]
                score = user_scores[edge]
                selected_items[u, pos] = item
                selected_scores[u, pos] = score
                counts[item] += 1
                total_score += score
                total_recs += 1

        diversity = 0
        for item in range(num_items):
            if counts[item] > 0:
                diversity += 1
        return selected_items, selected_scores, counts, total_score, total_recs, diversity


    @njit(cache=True)
    def _best_swap_for_item_numba(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts):
        n = selected_items.shape[1]
        best_loss = 1.0e30
        best_u = -1
        best_pos = -1
        best_add_score = 0.0
        best_remove_item = -1

        for p in range(item_indptr[item], item_indptr[item + 1]):
            u = item_users[p]
            add_score = item_scores[p]

            already_selected = False
            for pos in range(n):
                if selected_items[u, pos] == item:
                    already_selected = True
                    break
            if already_selected:
                continue

            for pos in range(n):
                remove_item = selected_items[u, pos]
                if remove_item < 0 or counts[remove_item] < 2:
                    continue
                loss = selected_scores[u, pos] - add_score
                if loss < best_loss:
                    best_loss = loss
                    best_u = u
                    best_pos = pos
                    best_add_score = add_score
                    best_remove_item = remove_item
        return best_loss, best_u, best_pos, best_add_score, best_remove_item


    @njit(cache=True)
    def _batch_greedy_history_numba(
        num_items,
        item_indptr,
        item_users,
        item_scores,
        selected_items,
        selected_scores,
        counts,
        initial_score,
        total_recs,
        initial_diversity,
        history_interval,
        max_rounds,
    ):
        max_hist = num_items + max_rounds + 4
        hist_iter = np.empty(max_hist, dtype=np.int64)
        hist_d = np.empty(max_hist, dtype=np.int64)
        hist_obj = np.empty(max_hist, dtype=np.float64)

        hist_len = 1
        hist_iter[0] = 0
        hist_d[0] = initial_diversity
        hist_obj[0] = initial_score

        total_score = initial_score
        diversity = initial_diversity
        swaps = 0
        rounds = 0
        best_loss_by_item = np.empty(num_items, dtype=np.float64)

        while rounds < max_rounds and diversity < num_items:
            rounds += 1
            for item in range(num_items):
                if counts[item] > 0:
                    best_loss_by_item[item] = 1.0e30
                else:
                    loss, u, pos, add_score, remove_item = _best_swap_for_item_numba(
                        item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts
                    )
                    if u >= 0:
                        best_loss_by_item[item] = loss
                    else:
                        best_loss_by_item[item] = 1.0e30

            order = np.argsort(best_loss_by_item)
            changed = 0
            for idx in range(num_items):
                item = order[idx]
                if best_loss_by_item[item] >= 1.0e29:
                    break
                if counts[item] > 0:
                    continue

                loss, u, pos, add_score, remove_item = _best_swap_for_item_numba(
                    item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts
                )
                if u < 0 or counts[item] > 0 or remove_item < 0 or counts[remove_item] < 2:
                    continue

                counts[item] += 1
                counts[remove_item] -= 1
                total_score += add_score - selected_scores[u, pos]
                selected_items[u, pos] = item
                selected_scores[u, pos] = add_score
                diversity += 1
                swaps += 1
                changed += 1

                if history_interval <= 1 or swaps % history_interval == 0:
                    hist_iter[hist_len] = swaps
                    hist_d[hist_len] = diversity
                    hist_obj[hist_len] = total_score
                    hist_len += 1
                if diversity >= num_items:
                    break

            if changed == 0:
                break

        if hist_iter[hist_len - 1] != swaps:
            hist_iter[hist_len] = swaps
            hist_d[hist_len] = diversity
            hist_obj[hist_len] = total_score
            hist_len += 1

        pred = total_score / total_recs if total_recs > 0 else 0.0
        return pred, diversity, swaps, rounds, total_score, hist_iter[:hist_len], hist_d[:hist_len], hist_obj[:hist_len]


def dataset_from_filename() -> str:
    name = Path(__file__).stem.upper()
    if name.startswith("YELP"):
        return "Yelp"
    for dataset in ("OP", "VG", "TG", "SO"):
        if name.startswith(dataset):
            return dataset
    return "OP"


def load_candidate_arrays(scores: np.ndarray, cand_path: Path):
    z = np.load(cand_path)
    user_indptr = z["user_indptr"].astype(np.int64, copy=False)
    user_items_key = "user_items" if "user_items" in z.files else "user_indices"
    user_items = z[user_items_key].astype(np.int32, copy=False)
    item_indptr = z["item_indptr"].astype(np.int64, copy=False)
    item_users = z["item_users"].astype(np.int32, copy=False)

    if "user_scores" in z.files:
        user_scores = z["user_scores"].astype(np.float32, copy=False)
    else:
        user_scores = np.empty(len(user_items), dtype=np.float32)
        for u in range(scores.shape[0]):
            s, e = int(user_indptr[u]), int(user_indptr[u + 1])
            if e > s:
                user_scores[s:e] = scores[u, user_items[s:e]]

    if "item_scores" in z.files:
        item_scores = z["item_scores"].astype(np.float32, copy=False)
    else:
        item_scores = np.empty(len(item_users), dtype=np.float32)
        for item in range(scores.shape[1]):
            s, e = int(item_indptr[item]), int(item_indptr[item + 1])
            if e > s:
                item_scores[s:e] = scores[item_users[s:e], item]

    return user_indptr, user_items, user_scores, item_indptr, item_users, item_scores


def first_reached(hist_d: np.ndarray, target: int) -> int:
    for i, value in enumerate(hist_d):
        if int(value) >= target:
            return i
    return len(hist_d) - 1


def format_workbook(path: Path):
    widths = {
        "迭代明细": {"A": 10, "B": 13, "C": 13, "D": 19, "E": 13},
        "D比例结果": {"A": 10, "B": 13, "C": 13, "D": 13, "E": 13, "F": 19, "G": 13, "H": 13, "I": 17},
        "总结果": {
            "A": 10,
            "B": 13,
            "C": 19,
            "D": 13,
            "E": 13,
            "F": 10,
            "G": 19,
            "H": 13,
            "I": 12,
            "J": 10,
            "K": 13,
            "L": 17,
            "M": 13,
            "N": 13,
        },
        "meta": {
            "A": 10,
            "B": 13,
            "C": 11,
            "D": 13,
            "E": 17,
            "F": 23,
            "G": 24,
            "H": 20,
            "I": 10,
            "J": 24,
            "K": 13,
            "L": 13,
            "M": 17,
            "N": 17,
        },
    }
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="宋体", size=11, bold=(cell.row == 1))
        for col, width in widths.get(ws.title, {}).items():
            ws.column_dimensions[col].width = width
    wb.save(path)


def write_outputs(out_xlsx: Path, out_json: Path, history_rows, percent_rows, summary_rows, meta_rows):
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        pd.DataFrame(history_rows)[["数据集", "迭代次数", "总体多样性", "目标函数值", "时间s"]].to_excel(
            writer, sheet_name="迭代明细", index=False
        )
        pd.DataFrame(percent_rows)[
            ["数据集", "D比例", "D_target", "迭代次数", "总体多样性", "目标函数值", "准确多样性", "时间s", "target_feasible"]
        ].to_excel(writer, sheet_name="D比例结果", index=False)
        pd.DataFrame(summary_rows)[
            [
                "数据集",
                "N",
                "总时间s",
                "目标函数值",
                "准确多样性",
                "总体多样性",
                "Naive目标函数值",
                "Naive准确多样性",
                "Naive总体多样性",
                "迭代次数",
                "交换次数",
                "target_feasible",
                "max_reached",
                "total_recs",
            ]
        ].to_excel(writer, sheet_name="总结果", index=False)
        pd.DataFrame(meta_rows).to_excel(writer, sheet_name="meta", index=False)

    format_workbook(out_xlsx)
    out_json.write_text(
        json.dumps(
            {"history": history_rows, "percent_results": percent_rows, "summary": summary_rows, "meta": meta_rows},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def run_dataset(dataset: str, n: int, history_interval: int, max_rounds: int):
    if njit is None:
        raise RuntimeError("当前环境没有 numba，贪心算法会非常慢。请先安装 numba 后再运行。")
    cfg = DATASETS[dataset]
    scores_path = Path(cfg["scores"])
    cand_path = Path(cfg["candidates"])
    out_dir = SCRIPT_DIR / f"{dataset}贪心算法_results"
    out_xlsx = out_dir / f"{dataset}贪心算法_结果.xlsx"
    out_json = out_dir / f"{dataset}贪心算法_结果.json"

    print(f"=== {dataset} ===", flush=True)
    print(f"numba_enabled={njit is not None}", flush=True)
    print(f"N={n}", flush=True)
    print(f"history_interval={history_interval}", flush=True)
    print(f"scores={scores_path}", flush=True)
    print(f"candidates={cand_path}", flush=True)
    print(f"output_excel={out_xlsx}", flush=True)

    if not scores_path.exists():
        raise FileNotFoundError(f"找不到 score 缓存: {scores_path}")
    if not cand_path.exists():
        raise FileNotFoundError(f"找不到 candidate 缓存: {cand_path}")

    scores = np.load(scores_path, mmap_mode="r")
    num_users, num_items = scores.shape

    t_arrays = time.time()
    user_indptr, user_items, user_scores, item_indptr, item_users, item_scores = load_candidate_arrays(scores, cand_path)
    array_build_sec = time.time() - t_arrays
    candidate_edges = int(len(user_items))
    items_with_candidates = int(np.count_nonzero(np.diff(item_indptr) > 0))
    print(f"shape users={num_users}, items={num_items}", flush=True)
    print(f"candidate_edges={candidate_edges:,}, items_with_candidates={items_with_candidates}", flush=True)

    t0 = time.time()
    selected_items, selected_scores, counts, naive_obj, total_recs, naive_diversity = _initial_topn_numba(
        int(num_users), int(num_items), user_indptr, user_items, user_scores, int(n)
    )
    naive_time = time.time() - t0
    naive_pred = float(naive_obj) / total_recs if total_recs else 0.0

    t1 = time.time()
    pred, diversity, swaps, rounds, obj, hist_iter, hist_d, hist_obj = _batch_greedy_history_numba(
        int(num_items),
        item_indptr,
        item_users,
        item_scores,
        selected_items,
        selected_scores,
        counts,
        float(naive_obj),
        int(total_recs),
        int(naive_diversity),
        int(history_interval),
        int(max_rounds),
    )
    greedy_time = time.time() - t1
    total_time = naive_time + greedy_time

    if len(hist_iter) > 1 and hist_iter[-1] > 0:
        times = naive_time + (hist_iter.astype(np.float64) / float(hist_iter[-1])) * greedy_time
    else:
        times = np.array([naive_time], dtype=np.float64)

    history_rows = []
    for it, div, objective, elapsed in zip(hist_iter, hist_d, hist_obj, times):
        history_rows.append(
            {
                "数据集": dataset,
                "迭代次数": int(it),
                "总体多样性": int(div),
                "目标函数值": float(objective),
                "时间s": float(elapsed),
            }
        )

    percent_rows = []
    for pct in D_PERCENTAGES:
        d_target = int(math.ceil(items_with_candidates * pct / 100.0))
        idx = first_reached(hist_d, d_target)
        percent_rows.append(
            {
                "数据集": dataset,
                "D比例": f"D-{pct}%",
                "D_target": d_target,
                "迭代次数": int(hist_iter[idx]),
                "总体多样性": int(hist_d[idx]),
                "目标函数值": float(hist_obj[idx]),
                "准确多样性": float(hist_obj[idx]) / total_recs if total_recs else 0.0,
                "时间s": float(times[idx]),
                "target_feasible": bool(int(hist_d[idx]) >= d_target),
            }
        )

    summary_rows = [
        {
            "数据集": dataset,
            "N": int(n),
            "总时间s": float(total_time),
            "目标函数值": float(obj),
            "准确多样性": float(pred),
            "总体多样性": int(diversity),
            "Naive目标函数值": float(naive_obj),
            "Naive准确多样性": float(naive_pred),
            "Naive总体多样性": int(naive_diversity),
            "迭代次数": int(swaps),
            "交换次数": int(swaps),
            "target_feasible": bool(diversity >= num_items),
            "max_reached": int(diversity),
            "total_recs": int(total_recs),
        }
    ]
    meta_rows = [
        {
            "dataset": dataset,
            "N": int(n),
            "num_users": int(num_users),
            "num_items": int(num_items),
            "candidate_edges": candidate_edges,
            "items_with_candidates": items_with_candidates,
            "D_percentages": str(D_PERCENTAGES),
            "candidate_fraction": cfg["candidate_fraction"],
            "seed": SEED,
            "engine": "numba_batch_greedy_algorithm2",
            "data_path": str(cfg["data_path"]),
            "cache_scores": str(scores_path),
            "cache_candidates": str(cand_path),
            "array_build_sec": float(array_build_sec),
            "batch_rounds": int(rounds),
        }
    ]

    write_outputs(out_xlsx, out_json, history_rows, percent_rows, summary_rows, meta_rows)
    print(summary_rows[0], flush=True)
    print(f"saved_excel={out_xlsx}", flush=True)
    print(f"saved_json={out_json}", flush=True)


def parse_args():
    parser = argparse.ArgumentParser(description="五个数据集统一批量贪心算法脚本，默认按文件名选择数据集。")
    parser.add_argument("--dataset", choices=["OP", "Yelp", "VG", "TG", "SO"], default=dataset_from_filename())
    parser.add_argument("--n", type=int, default=DEFAULT_N)
    parser.add_argument("--history-interval", type=int, default=DEFAULT_HISTORY_INTERVAL)
    parser.add_argument("--max-rounds", type=int, default=DEFAULT_MAX_ROUNDS)
    return parser.parse_args()


def main():
    args = parse_args()
    if args.n <= 0:
        raise ValueError("--n 必须大于 0")
    if args.history_interval <= 0:
        raise ValueError("--history-interval 必须大于 0")
    if args.max_rounds <= 0:
        raise ValueError("--max-rounds 必须大于 0")
    run_dataset(args.dataset, args.n, args.history_interval, args.max_rounds)


if __name__ == "__main__":
    main()
