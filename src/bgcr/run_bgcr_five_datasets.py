
from __future__ import annotations

import argparse
import json
import math
import time
import zipfile
import xml.etree.ElementTree as ET
try:
    from lxml import etree as LET
except Exception:
    LET = None
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

try:
    from numba import njit
except Exception:
    njit = None


ROOT = Path(r"C:\Users\24qxh\Documents\Codex\2026-07-04\new-chat")
CACHE_DIR = Path(r"E:\Users\24qxh\Desktop\greedy_cache")
OUT_DIR = ROOT / "outputs" / "greedy_5datasets_history"
OUT_XLSX = OUT_DIR / "five_datasets_greedy_summary_and_history.xlsx"
OUT_JSON = OUT_DIR / "five_datasets_greedy_summary_and_history.json"
FIG_DIR = OUT_DIR / "figures"
YELP_EXISTING_CACHE = ROOT / "work" / "yelp_cache"

N_VALUES = [10, 15, 20, 25, 30]
CANDIDATE_FRACTION = 0.40
SEED = 20260704
GREEDY_PASSES = 2
HISTORY_INTERVAL = 50

DATASETS = {
    "OP": [Path(r"E:\Users\24qxh\Desktop\op\IUIC1_Rui_pred_user_item_UNKNOWN_ONLY_known_ratings_0_item_headers(1).xlsx")],
    "VG": [Path(r"E:\Users\24qxh\Desktop\VG\VG_R_ui_UNKNOWN_ONLY_known_ratings_0_display1_1.xlsx")],
    "Yelp": [Path(r"E:\Users\24qxh\Desktop\Yelp_R_ui_UNKNOWN_ONLY_known_ratings_0_display1.xlsx")],
    "SO": [
        Path(r"E:\Users\24qxh\Desktop\推荐系统\TG_reproduce\SO_unobserved_R_ui_part1.csv"),
        Path(r"E:\Users\24qxh\Desktop\推荐系统\TG_reproduce\SO_unobserved_R_ui_part2.csv"),
        Path(r"E:\Users\24qxh\Desktop\推荐系统\TG_reproduce\SO_unobserved_R_ui_part3.csv"),
    ],
    "TG": [Path(r"E:\Users\24qxh\Desktop\TG_R_ui_UNKNOWN_ONLY_known_ratings_0.xlsx")],
}


def safe_name(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in name)


def is_id_column(values: pd.Series) -> bool:
    numeric = pd.to_numeric(values, errors="coerce")
    return numeric.isna().mean() > 0.5


def excel_col_to_int(cell_ref: str) -> int:
    value = 0
    for ch in cell_ref:
        if not ch.isalpha():
            break
        value = value * 26 + (ord(ch.upper()) - ord("A") + 1)
    return value


def parse_xlsx_dimension(ref: str):
    if not ref:
        return None
    last = ref.split(":")[-1]
    letters = "".join(ch for ch in last if ch.isalpha())
    digits = "".join(ch for ch in last if ch.isdigit())
    if not letters or not digits:
        return None
    return int(digits), excel_col_to_int(letters)


def load_xlsx_matrix_fast_xml(path: Path) -> np.ndarray:
    with zipfile.ZipFile(path) as zf:
        sheet_names = sorted(
            name for name in zf.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if not sheet_names:
            raise ValueError("no worksheet xml found")
        sheet_name = sheet_names[0]

        n_rows = None
        n_cols_total = None
        with zf.open(sheet_name) as f:
            for event, elem in ET.iterparse(f, events=("start",)):
                if elem.tag.endswith("dimension"):
                    dim = parse_xlsx_dimension(elem.attrib.get("ref", ""))
                    if dim is not None:
                        n_rows, n_cols_total = dim
                    break

        if n_rows is None or n_cols_total is None or n_rows <= 1 or n_cols_total <= 1:
            raise ValueError("cannot infer worksheet dimension")

        arr = np.zeros((n_rows - 1, n_cols_total - 1), dtype=np.float32)
        with zf.open(sheet_name) as f:
            parser = LET.iterparse(f, events=("end",), recover=True, huge_tree=True) if LET is not None else ET.iterparse(f, events=("end",))
            for event, elem in parser:
                if not elem.tag.endswith("row"):
                    continue
                row_ref = elem.attrib.get("r")
                if row_ref is None:
                    elem.clear()
                    continue
                row_idx = int(row_ref)
                if row_idx <= 1 or row_idx > n_rows:
                    elem.clear()
                    continue
                out_row = row_idx - 2
                col_seq = 0
                for cell in elem:
                    if not cell.tag.endswith("c"):
                        continue
                    col_seq += 1
                    ref = cell.attrib.get("r", "")
                    col = excel_col_to_int(ref) if ref else col_seq
                    if col <= 1 or col > n_cols_total:
                        continue
                    v_text = None
                    for child in cell:
                        if child.tag.endswith("v"):
                            v_text = child.text
                            break
                    if v_text:
                        try:
                            arr[out_row, col - 2] = float(v_text)
                        except ValueError:
                            pass
                elem.clear()
        return arr


def load_xlsx_matrix(path: Path) -> np.ndarray:
    try:
        print("  reading xlsx by fast xml stream", flush=True)
        return load_xlsx_matrix_fast_xml(path)
    except Exception as exc:
        print(f"  fast xml read failed: {exc}", flush=True)

    try:
        df = pd.read_excel(path, engine="calamine")
        if len(df.columns) > 1:
            df = df.iloc[:, 1:]
        return df.apply(pd.to_numeric, errors="coerce").fillna(0.0).to_numpy(dtype=np.float32)
    except Exception as exc:
        print(f"  calamine read failed, fallback to openpyxl: {exc}", flush=True)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        n_cols = len(header) - 1
        data = []
        for row in rows:
            if row is None or row[0] is None:
                continue
            data.append([0.0 if x is None else float(x) for x in row[1:1 + n_cols]])
    finally:
        wb.close()
    return np.asarray(data, dtype=np.float32)


def load_csv_matrix(path: Path) -> np.ndarray:
    df = pd.read_csv(path)
    if len(df.columns) > 1 and is_id_column(df.iloc[:, 0]):
        df = df.iloc[:, 1:]
    return df.apply(pd.to_numeric, errors="coerce").fillna(0.0).to_numpy(dtype=np.float32)


def load_dataset_matrix(name: str, paths: Iterable[Path]) -> np.ndarray:
    if name == "Yelp":
        yelp_cache = YELP_EXISTING_CACHE / "yelp_unknown_scores_float32.npy"
        if yelp_cache.exists():
            print(f"  loaded existing Yelp cache: {yelp_cache}", flush=True)
            return np.load(yelp_cache, mmap_mode="r")

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    paths = list(paths)
    cache = CACHE_DIR / f"{safe_name(name)}_scores_float32.npy"
    meta = CACHE_DIR / f"{safe_name(name)}_scores_meta.json"
    newest = max(p.stat().st_mtime for p in paths)
    if cache.exists() and meta.exists():
        info = json.loads(meta.read_text(encoding="utf-8"))
        if info.get("newest_mtime", 0) >= newest:
            print(f"  loaded score cache: {cache}", flush=True)
            return np.load(cache, mmap_mode="r")

    parts = []
    for path in paths:
        print(f"  reading {path}", flush=True)
        if path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}:
            part = load_xlsx_matrix(path)
        elif path.suffix.lower() == ".csv":
            part = load_csv_matrix(path)
        else:
            raise ValueError(f"Unsupported file type: {path}")
        parts.append(part)

    rows = {part.shape[0] for part in parts}
    if len(rows) != 1:
        raise ValueError(f"{name} parts have different row counts: {[p.shape for p in parts]}")
    scores = parts[0] if len(parts) == 1 else np.concatenate(parts, axis=1)
    np.save(cache, scores.astype(np.float32, copy=False))
    meta.write_text(json.dumps({"shape": scores.shape, "newest_mtime": newest}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  cached scores: {cache}, shape={scores.shape}", flush=True)
    return np.load(cache, mmap_mode="r")


def build_candidate_csr(name: str, scores: np.ndarray):
    if name == "Yelp":
        compat_cache = CACHE_DIR / f"{safe_name(name)}_cand_frac_0p4_seed{SEED}.npz"
        if compat_cache.exists():
            zc = np.load(compat_cache)
            print(f"  loaded candidate cache: {compat_cache}", flush=True)
            return tuple(zc[k] for k in ["user_indptr", "user_items", "user_scores", "item_indptr", "item_users", "item_scores"])

        yelp_cand = YELP_EXISTING_CACHE / f"yelp_cand_frac_0p4_seed{SEED}.npz"
        if yelp_cand.exists():
            z = np.load(yelp_cand)
            print(f"  loaded existing Yelp candidate cache: {yelp_cand}", flush=True)
            user_indptr = z["user_indptr"].astype(np.int64, copy=False)
            user_items = z["user_indices"].astype(np.int32, copy=False)
            item_indptr = z["item_indptr"].astype(np.int64, copy=False)
            item_users = z["item_users"].astype(np.int32, copy=False)
            user_scores = np.empty(len(user_items), dtype=np.float32)
            for u in range(scores.shape[0]):
                s, e = int(user_indptr[u]), int(user_indptr[u + 1])
                if e > s:
                    user_scores[s:e] = scores[u, user_items[s:e]]
            item_scores = np.empty(len(item_users), dtype=np.float32)
            for item in range(scores.shape[1]):
                s, e = int(item_indptr[item]), int(item_indptr[item + 1])
                if e > s:
                    item_scores[s:e] = scores[item_users[s:e], item]
            np.savez(compat_cache, user_indptr=user_indptr, user_items=user_items, user_scores=user_scores, item_indptr=item_indptr, item_users=item_users, item_scores=item_scores)
            print(f"  cached compatible Yelp candidates: {compat_cache}", flush=True)
            return user_indptr, user_items, user_scores, item_indptr, item_users, item_scores

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache = CACHE_DIR / f"{safe_name(name)}_cand_frac_0p4_seed{SEED}.npz"
    if cache.exists():
        z = np.load(cache)
        print(f"  loaded candidate cache: {cache}", flush=True)
        return tuple(z[k] for k in ["user_indptr", "user_items", "user_scores", "item_indptr", "item_users", "item_scores"])

    rng = np.random.default_rng(SEED)
    num_users, num_items = scores.shape
    user_indptr = [0]
    user_items = []
    user_scores = []
    item_users = [[] for _ in range(num_items)]
    item_scores = [[] for _ in range(num_items)]

    for u in range(num_users):
        row = np.asarray(scores[u], dtype=np.float32)
        positive = np.flatnonzero(np.isfinite(row) & (row > 0.0))
        if positive.size == 0:
            user_indptr.append(len(user_items))
            continue
        k = max(1, int(math.ceil(CANDIDATE_FRACTION * positive.size)))
        jitter = rng.random(positive.size, dtype=np.float32) * np.float32(1e-7)
        local = np.argpartition(row[positive] + jitter, -k)[-k:]
        cand = np.sort(positive[local].astype(np.int32, copy=False))
        for item in cand:
            score = float(row[int(item)])
            user_items.append(int(item))
            user_scores.append(score)
            item_users[int(item)].append(u)
            item_scores[int(item)].append(score)
        user_indptr.append(len(user_items))
        if (u + 1) % 1000 == 0:
            print(f"  candidate users {u + 1}/{num_users}", flush=True)

    item_indptr = [0]
    flat_item_users = []
    flat_item_scores = []
    for users, vals in zip(item_users, item_scores):
        flat_item_users.extend(users)
        flat_item_scores.extend(vals)
        item_indptr.append(len(flat_item_users))

    arrays = (
        np.asarray(user_indptr, dtype=np.int64),
        np.asarray(user_items, dtype=np.int32),
        np.asarray(user_scores, dtype=np.float32),
        np.asarray(item_indptr, dtype=np.int64),
        np.asarray(flat_item_users, dtype=np.int32),
        np.asarray(flat_item_scores, dtype=np.float32),
    )
    np.savez(cache, user_indptr=arrays[0], user_items=arrays[1], user_scores=arrays[2], item_indptr=arrays[3], item_users=arrays[4], item_scores=arrays[5])
    print(f"  cached candidates: {cache}", flush=True)
    return arrays


def initial_topn(num_users: int, num_items: int, user_indptr, user_items, user_scores, n: int):
    selected_items = np.full((num_users, n), -1, dtype=np.int32)
    selected_scores = np.zeros((num_users, n), dtype=np.float32)
    counts = np.zeros(num_items, dtype=np.int32)
    total_score = 0.0
    total_recs = 0

    for u in range(num_users):
        s, e = int(user_indptr[u]), int(user_indptr[u + 1])
        length = e - s
        if length <= 0:
            continue
        k = min(n, length)
        vals = user_scores[s:e]
        local = np.argpartition(vals, -k)[-k:]
        local = local[np.argsort(-vals[local])]
        for pos, loc in enumerate(local):
            edge = s + int(loc)
            item = int(user_items[edge])
            score = float(user_scores[edge])
            selected_items[u, pos] = item
            selected_scores[u, pos] = score
            counts[item] += 1
            total_score += score
            total_recs += 1
    return selected_items, selected_scores, counts, total_score, total_recs


if njit is not None:
    @njit(cache=True)
    def best_swap_for_item_numba(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts):
        n = selected_items.shape[1]
        best_loss = 1.0e30
        best_u = -1
        best_pos = -1
        best_add_score = 0.0
        best_remove_item = -1
        for p in range(item_indptr[item], item_indptr[item + 1]):
            u = item_users[p]
            add_score = item_scores[p]
            already = False
            for pos in range(n):
                if selected_items[u, pos] == item:
                    already = True
                    break
            if already:
                continue
            for pos in range(n):
                rem_item = selected_items[u, pos]
                if rem_item < 0 or counts[rem_item] < 2:
                    continue
                loss = selected_scores[u, pos] - add_score
                if loss < best_loss:
                    best_loss = loss
                    best_u = u
                    best_pos = pos
                    best_add_score = add_score
                    best_remove_item = rem_item
        return best_loss, best_u, best_pos, best_add_score, best_remove_item


    @njit(cache=True)
    def greedy_repair_numba(num_items, item_indptr, item_users, item_scores, selected_items, selected_scores, counts, total_score, total_recs, passes, interval):
        swaps = 0
        d = 0
        for item in range(num_items):
            if counts[item] > 0:
                d += 1

        max_hist = num_items + 2
        hist_iter = np.empty(max_hist, dtype=np.int64)
        hist_d = np.empty(max_hist, dtype=np.int64)
        hist_obj = np.empty(max_hist, dtype=np.float64)
        hist_len = 0
        hist_iter[hist_len] = 0
        hist_d[hist_len] = d
        hist_obj[hist_len] = total_score
        hist_len += 1

        best_loss = np.empty(num_items, dtype=np.float32)
        for _pass in range(passes):
            for item in range(num_items):
                if counts[item] > 0:
                    best_loss[item] = np.inf
                else:
                    loss, u, pos, add_score, rem_item = best_swap_for_item_numba(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts)
                    best_loss[item] = loss if u >= 0 else np.inf

            order = np.argsort(best_loss)
            changed = 0
            for idx in range(num_items):
                item = order[idx]
                if not np.isfinite(best_loss[item]) or counts[item] > 0:
                    continue
                loss, u, pos, add_score, rem_item = best_swap_for_item_numba(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts)
                if u < 0 or counts[item] > 0 or rem_item < 0 or counts[rem_item] < 2:
                    continue
                counts[item] += 1
                counts[rem_item] -= 1
                d += 1
                total_score += add_score - selected_scores[u, pos]
                selected_items[u, pos] = item
                selected_scores[u, pos] = add_score
                swaps += 1
                changed += 1
                if swaps % interval == 0 and hist_len < max_hist:
                    hist_iter[hist_len] = swaps
                    hist_d[hist_len] = d
                    hist_obj[hist_len] = total_score
                    hist_len += 1
            if changed == 0:
                break

        if hist_len == 0 or hist_iter[hist_len - 1] != swaps:
            hist_iter[hist_len] = swaps
            hist_d[hist_len] = d
            hist_obj[hist_len] = total_score
            hist_len += 1
        pred = total_score / total_recs if total_recs > 0 else 0.0
        return pred, d, swaps, total_score, hist_iter[:hist_len], hist_d[:hist_len], hist_obj[:hist_len]


def best_swap_for_item_python(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts):
    n = selected_items.shape[1]
    best_loss = math.inf
    best = None
    for p in range(int(item_indptr[item]), int(item_indptr[item + 1])):
        u = int(item_users[p])
        add_score = float(item_scores[p])
        if np.any(selected_items[u] == item):
            continue
        for pos in range(n):
            rem_item = int(selected_items[u, pos])
            if rem_item < 0 or counts[rem_item] < 2:
                continue
            loss = float(selected_scores[u, pos]) - add_score
            if loss < best_loss:
                best_loss = loss
                best = (item, u, pos, add_score, rem_item)
    return best_loss, best


def greedy_repair(num_items, item_indptr, item_users, item_scores, selected_items, selected_scores, counts, total_score, total_recs):
    if njit is not None:
        pred, d, swaps, obj, hi, hd, ho = greedy_repair_numba(
            int(num_items),
            item_indptr.astype(np.int64, copy=False),
            item_users.astype(np.int32, copy=False),
            item_scores.astype(np.float32, copy=False),
            selected_items,
            selected_scores,
            counts,
            float(total_score),
            int(total_recs),
            int(GREEDY_PASSES),
            int(HISTORY_INTERVAL),
        )
        history = [(int(hi[i]), int(hd[i]), float(ho[i])) for i in range(len(hi))]
        return float(pred), int(d), int(swaps), float(obj), history

    swaps = 0
    d = int(np.count_nonzero(counts))
    history = [(0, d, float(total_score))]
    for _pass in range(GREEDY_PASSES):
        candidates = []
        for item in range(num_items):
            if counts[item] > 0:
                continue
            loss, best = best_swap_for_item_python(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts)
            if best is not None:
                candidates.append((loss, item))
        changed = 0
        for _, item in sorted(candidates):
            if counts[item] > 0:
                continue
            _, best = best_swap_for_item_python(item, item_indptr, item_users, item_scores, selected_items, selected_scores, counts)
            if best is None:
                continue
            item, u, pos, add_score, rem_item = best
            if counts[rem_item] < 2:
                continue
            counts[item] += 1
            counts[rem_item] -= 1
            d += 1
            total_score += add_score - float(selected_scores[u, pos])
            selected_items[u, pos] = item
            selected_scores[u, pos] = add_score
            swaps += 1
            changed += 1
            if swaps % HISTORY_INTERVAL == 0:
                history.append((swaps, d, float(total_score)))
        if changed == 0:
            break
    if history[-1][0] != swaps:
        history.append((swaps, d, float(total_score)))
    pred = total_score / total_recs if total_recs else 0.0
    return pred, d, swaps, float(total_score), history


def run_dataset(name: str, paths: Iterable[Path]):
    print(f"=== {name} ===", flush=True)
    scores = load_dataset_matrix(name, paths)
    num_users, num_items = scores.shape
    user_indptr, user_items, user_scores, item_indptr, item_users, item_scores = build_candidate_csr(name, scores)
    rows = []
    history_rows = []

    for n in N_VALUES:
        print(f"--- {name}, N={n} ---", flush=True)
        t0 = time.time()
        selected_items, selected_scores, counts, naive_obj, total_recs = initial_topn(num_users, num_items, user_indptr, user_items, user_scores, n)
        naive_pred = naive_obj / total_recs if total_recs else 0.0
        naive_d = int(np.count_nonzero(counts))
        greedy_pred, greedy_d, swaps, greedy_obj, hist = greedy_repair(num_items, item_indptr, item_users, item_scores, selected_items, selected_scores, counts, naive_obj, total_recs)
        elapsed = time.time() - t0
        row = {
            "数据集": name,
            "N": n,
            "Naive Pred": naive_pred,
            "Naive D": naive_d,
            "Naive目标函数值": naive_obj,
            "Greedy Pred": greedy_pred,
            "Greedy D": greedy_d,
            "Greedy目标函数值": greedy_obj,
            "准确率损失%": (greedy_pred - naive_pred) / naive_pred * 100 if naive_pred else 0.0,
            "覆盖提升%": (greedy_d - naive_d) / naive_d * 100 if naive_d else 0.0,
            "时间s": elapsed,
            "迭代次数": swaps,
            "用户数": num_users,
            "项目数": num_items,
            "总推荐数": total_recs,
        }
        rows.append(row)
        for it, d, obj in hist:
            history_rows.append({
                "数据集": name,
                "N": n,
                "迭代次数": it,
                "总体多样性": d,
                "目标函数值": obj,
            })
        print(row, flush=True)
    return rows, history_rows


def write_excel(summary_rows, history_rows):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    detail = pd.DataFrame(summary_rows)
    history = pd.DataFrame(history_rows)
    paper = pd.DataFrame({
        "数据集": detail["数据集"],
        "N 取值": detail["N"],
        "推荐准确性-前人模型": detail["Naive Pred"],
        "推荐准确性-本文模型": detail["Greedy Pred"],
        "推荐准确性-提升比/%": detail["准确率损失%"],
        "总体多样性-前人模型": detail["Naive D"],
        "总体多样性-本文模型": detail["Greedy D"],
        "总体多样性-提升/%": detail["覆盖提升%"],
    })
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        paper.to_excel(writer, sheet_name="论文表_贪心为本文模型", index=False)
        detail.to_excel(writer, sheet_name="汇总指标含时间目标函数", index=False)
        history.to_excel(writer, sheet_name="迭代明细", index=False)

    wb = load_workbook(OUT_XLSX)
    thin = Side(style="thin", color="000000")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 24)
    wb.save(OUT_XLSX)


def write_figures(history_rows):
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    history = pd.DataFrame(history_rows)
    for dataset, df_d in history.groupby("数据集"):
        fig, axes = plt.subplots(1, 2, figsize=(12, 4), dpi=160)
        for n, df_n in df_d.groupby("N"):
            df_n = df_n.sort_values("迭代次数")
            axes[0].plot(df_n["迭代次数"], df_n["总体多样性"], label=f"N={n}")
            axes[1].plot(df_n["迭代次数"], df_n["目标函数值"], label=f"N={n}")
        axes[0].set_title(f"{dataset} Diversity")
        axes[0].set_xlabel("Iteration")
        axes[0].set_ylabel("Diversity D")
        axes[1].set_title(f"{dataset} Objective")
        axes[1].set_xlabel("Iteration")
        axes[1].set_ylabel("Objective value")
        for ax in axes:
            ax.grid(True, alpha=0.25)
            ax.legend(fontsize=8)
        fig.tight_layout()
        path = FIG_DIR / f"{safe_name(dataset)}_greedy_iteration_curves.png"
        fig.savefig(path)
        plt.close(fig)


def main():
    global HISTORY_INTERVAL
    parser = argparse.ArgumentParser()
    parser.add_argument("--datasets", nargs="*", default=list(DATASETS.keys()), help="Datasets to run, e.g. OP Yelp VG SO TG")
    parser.add_argument("--history-interval", type=int, default=HISTORY_INTERVAL)
    args = parser.parse_args()

    HISTORY_INTERVAL = int(args.history_interval)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    all_summary = []
    all_history = []
    if OUT_JSON.exists():
        try:
            old = json.loads(OUT_JSON.read_text(encoding="utf-8"))
            all_summary = old.get("summary", [])
            all_history = old.get("history", [])
        except Exception:
            pass

    existing = {(r.get("数据集"), r.get("N")) for r in all_summary}
    for name in args.datasets:
        if name not in DATASETS:
            raise ValueError(f"Unknown dataset: {name}")
        summary, history = run_dataset(name, DATASETS[name])
        for row in summary:
            key = (row.get("数据集"), row.get("N"))
            all_summary = [r for r in all_summary if (r.get("数据集"), r.get("N")) != key]
            all_summary.append(row)
        all_history = [r for r in all_history if r.get("数据集") != name]
        all_history.extend(history)
        OUT_JSON.write_text(json.dumps({"summary": all_summary, "history": all_history}, ensure_ascii=False, indent=2), encoding="utf-8")
        write_excel(all_summary, all_history)
        write_figures(all_history)
    print(OUT_XLSX, flush=True)
    print(OUT_JSON, flush=True)
    print(FIG_DIR, flush=True)


if __name__ == "__main__":
    main()
