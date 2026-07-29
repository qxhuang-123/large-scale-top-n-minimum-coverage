from __future__ import annotations

import json
import math
import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

try:
    import matplotlib.pyplot as plt
except ImportError as exc:
    raise SystemExit("Please install matplotlib first: pip install matplotlib") from exc


# =========================
# User settings
# =========================

# Top-N recommendation length.
N = 10

# D is fixed to I, namely the total number of items in each dataset.
D_EQUALS_I = True

# Risk coefficient alpha: 0.1, 0.2, ..., 1.0.
ALPHAS = [round(i / 10, 1) for i in range(1, 11)]

# The compiled exact-algorithm module used by your existing OP/Yelp runners.
# If PyCharm cannot import tpcar_core_fast, set this path to the folder that
# contains tpcar_core_fast*.pyd.
CPP_DIR = Path(
    r"C:\Users\24qxh\Documents\Codex\2026-07-05\import-json-import-sys-import-time-4\work\tpcar_cpp39_history"
)

# Input files and cache folders.
OP_XLSX = Path(r"E:\Users\24qxh\Desktop\op\IUIC1_Rui_pred_user_item_FULL_display1.xlsx")
YELP_XLSX = Path(r"E:\Users\24qxh\Desktop\Yelp\Yelp_R_ui_UNKNOWN_ONLY_known_ratings_0_display1.xlsx")

# These caches avoid reading very large Excel files repeatedly.
OUT_DIR = Path(r"E:\Users\24qxh\Desktop\exact_alpha_sensitivity")
CACHE_DIR = OUT_DIR / "cache"

OP_SCORE_CACHE = CACHE_DIR / "op_scores_float32.npy"
YELP_SCORE_CACHE = CACHE_DIR / "yelp_unknown_scores_float32.npy"

# If you already have the old Yelp score cache from previous runs, the script
# will use it first because reading the Yelp Excel file is slow.
EXISTING_YELP_SCORE_CACHE = Path(
    r"C:\Users\24qxh\Documents\Codex\2026-07-04\new-chat\work\yelp_cache\yelp_unknown_scores_float32.npy"
)

PROGRESS_EVERY = 200
SEED = 20260704


sys.path.insert(0, str(CPP_DIR))
try:
    import tpcar_core_fast as tpcar_core
except ImportError as exc:
    raise SystemExit(
        "Cannot import tpcar_core_fast. Check CPP_DIR at the top of this file, "
        "and make sure PyCharm uses a Python version compatible with the .pyd file."
    ) from exc


def setup_matplotlib() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei",
        "SimHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def read_excel_matrix_to_cache(xlsx_path: Path, cache_path: Path) -> np.ndarray:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    meta_path = cache_path.with_suffix(".meta.json")
    source_mtime = xlsx_path.stat().st_mtime

    if cache_path.exists() and meta_path.exists():
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        if meta.get("source_mtime", 0) >= source_mtime:
            print(f"Loaded score cache: {cache_path}", flush=True)
            return np.load(cache_path, mmap_mode="r")

    print(f"Reading Excel matrix: {xlsx_path}", flush=True)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        num_items = len(header) - 1
        data = []

        for row_idx, row in enumerate(rows, start=1):
            if row is None or row[0] is None:
                continue
            values = row[1 : 1 + num_items]
            data.append([0.0 if x is None or x == "" else float(x) for x in values])
            if row_idx % 500 == 0:
                print(f"  read {row_idx:,} users", flush=True)
    finally:
        wb.close()

    scores = np.asarray(data, dtype=np.float32)
    np.save(cache_path, scores)
    meta_path.write_text(
        json.dumps(
            {
                "source": str(xlsx_path),
                "source_mtime": source_mtime,
                "shape": list(scores.shape),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Cached scores: {cache_path}, shape={scores.shape}", flush=True)
    return np.load(cache_path, mmap_mode="r")


def load_scores(dataset: str) -> np.ndarray:
    if dataset == "OP":
        return read_excel_matrix_to_cache(OP_XLSX, OP_SCORE_CACHE)

    if dataset == "Yelp":
        if YELP_SCORE_CACHE.exists():
            print(f"Loaded score cache: {YELP_SCORE_CACHE}", flush=True)
            return np.load(YELP_SCORE_CACHE, mmap_mode="r")
        if EXISTING_YELP_SCORE_CACHE.exists():
            print(f"Using existing Yelp score cache: {EXISTING_YELP_SCORE_CACHE}", flush=True)
            return np.load(EXISTING_YELP_SCORE_CACHE, mmap_mode="r")
        return read_excel_matrix_to_cache(YELP_XLSX, YELP_SCORE_CACHE)

    raise ValueError(f"Unknown dataset: {dataset}")


def candidate_cache_path(dataset: str, alpha: float) -> Path:
    alpha_tag = f"{alpha:.1f}".replace(".", "p")
    return CACHE_DIR / f"{dataset.lower()}_cand_alpha_{alpha_tag}_seed{SEED}.npz"


def build_candidate_csr(scores: np.ndarray, dataset: str, alpha: float, n: int):
    cache_path = candidate_cache_path(dataset, alpha)
    if cache_path.exists():
        print(f"Loaded candidate cache: {cache_path}", flush=True)
        return np.load(cache_path)

    print(f"Building candidate CSR: dataset={dataset}, alpha={alpha:.1f}", flush=True)
    rng = np.random.default_rng(SEED)
    num_users, num_items = scores.shape
    user_indptr = [0]
    user_indices: list[int] = []
    item_users_lists: list[list[int]] = [[] for _ in range(num_items)]

    t0 = time.time()
    for u in range(num_users):
        row = np.asarray(scores[u], dtype=np.float32)
        valid = np.flatnonzero(np.isfinite(row) & (row > 0.0))
        if valid.size == 0:
            user_indptr.append(len(user_indices))
            continue

        k = max(n, int(math.ceil(alpha * valid.size)))
        k = min(k, valid.size)

        if k < valid.size:
            jitter = rng.random(valid.size, dtype=np.float32) * np.float32(1e-7)
            chosen = np.argpartition(row[valid] + jitter, -k)[-k:]
            items = valid[chosen]
        else:
            items = valid

        item_scores = row[items]
        order = np.argsort(-item_scores, kind="mergesort")
        items = items[order].astype(np.int32, copy=False)

        for item in items:
            item_int = int(item)
            user_indices.append(item_int)
            item_users_lists[item_int].append(u)
        user_indptr.append(len(user_indices))

        if (u + 1) % 500 == 0:
            print(f"  candidate users {u + 1:,}/{num_users:,}", flush=True)

    item_indptr = [0]
    item_users: list[int] = []
    for users in item_users_lists:
        item_users.extend(users)
        item_indptr.append(len(item_users))

    np.savez(
        cache_path,
        user_indptr=np.asarray(user_indptr, dtype=np.int32),
        user_indices=np.asarray(user_indices, dtype=np.int32),
        item_indptr=np.asarray(item_indptr, dtype=np.int32),
        item_users=np.asarray(item_users, dtype=np.int32),
    )
    print(
        f"Cached candidates: {cache_path}, edges={len(user_indices):,}, "
        f"time={time.time() - t0:.1f}s",
        flush=True,
    )
    return np.load(cache_path)


def build_cpp_arrays(scores: np.ndarray, z):
    user_indptr = z["user_indptr"].astype(np.int32, copy=False)
    user_items = z["user_indices"].astype(np.int32, copy=False)
    item_indptr = z["item_indptr"].astype(np.int32, copy=False)
    item_users = z["item_users"].astype(np.int32, copy=False)

    num_users, _ = scores.shape
    num_edges = len(user_items)
    edge_users = np.empty(num_edges, dtype=np.int32)
    user_scores = np.empty(num_edges, dtype=np.float64)

    for u in range(num_users):
        start, end = int(user_indptr[u]), int(user_indptr[u + 1])
        if end > start:
            items = user_items[start:end]
            edge_users[start:end] = u
            user_scores[start:end] = scores[u, items]

    item_edges = np.empty(num_edges, dtype=np.int32)
    cursor = item_indptr[:-1].copy()
    for edge_id, item in enumerate(user_items):
        pos = cursor[item]
        item_edges[pos] = edge_id
        cursor[item] += 1

    item_scores = user_scores[item_edges]
    return (
        user_indptr,
        user_items,
        user_scores,
        edge_users,
        item_indptr,
        item_users,
        item_edges,
        item_scores,
    )


def run_one_alpha(dataset: str, scores: np.ndarray, alpha: float) -> dict:
    num_users, num_items = scores.shape
    z = build_candidate_csr(scores, dataset, alpha, N)
    candidate_edges = int(len(z["user_indices"]))
    items_with_candidates = int((np.diff(z["item_indptr"]) > 0).sum())
    d_target = int(num_items) if D_EQUALS_I else int(items_with_candidates)

    print(
        f"--- EXACT {dataset}: N={N}, alpha={alpha:.1f}, D={d_target}, "
        f"items_with_candidates={items_with_candidates}/{num_items}, edges={candidate_edges:,} ---",
        flush=True,
    )

    t0 = time.time()
    arrays = build_cpp_arrays(scores, z)
    array_build_sec = time.time() - t0

    t1 = time.time()
    raw = dict(
        tpcar_core.run_exact_csr(
            int(num_users),
            int(num_items),
            *arrays,
            int(N),
            int(d_target),
            int(PROGRESS_EVERY),
        )
    )
    solve_sec = time.time() - t1

    total_recs = int(raw["total_recs"])
    accuracy = float(raw["tpcar_pred"])
    diversity = int(raw["tpcar_D"])

    row = {
        "dataset": dataset,
        "N": N,
        "alpha": alpha,
        "risk_coefficient_percent": int(round(alpha * 100)),
        "D_target": d_target,
        "num_users": int(num_users),
        "num_items": int(num_items),
        "candidate_edges": candidate_edges,
        "items_with_candidates": items_with_candidates,
        "total_recs": total_recs,
        "recommendation_accuracy": accuracy,
        "overall_diversity": diversity,
        "objective": float(raw["tpcar_obj"]),
        "naive_accuracy": float(raw["naive_pred"]),
        "naive_diversity": int(raw["naive_D"]),
        "augmentations": int(raw["augmentations"]),
        "swaps": int(raw["swaps"]),
        "target_feasible": bool(raw["target_feasible"]),
        "max_reached": int(raw["max_reached"]),
        "array_build_sec": round(array_build_sec, 3),
        "solve_sec": round(solve_sec, 3),
        "total_sec": round(array_build_sec + solve_sec, 3),
    }
    print(
        f"{dataset} alpha={alpha:.1f}: accuracy={accuracy:.6f}, "
        f"diversity={diversity}, time={row['total_sec']}s",
        flush=True,
    )
    return row


def write_outputs(results: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(results)
    csv_path = OUT_DIR / "exact_alpha_sensitivity_results.csv"
    xlsx_path = OUT_DIR / "exact_alpha_sensitivity_results.xlsx"
    acc_png = OUT_DIR / "exact_alpha_sensitivity_accuracy.png"
    div_png = OUT_DIR / "exact_alpha_sensitivity_diversity.png"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="alpha_sensitivity", index=False)

    setup_matplotlib()
    plot_metric(
        df,
        metric="recommendation_accuracy",
        ylabel="推荐准确性",
        title="风险系数 alpha 对推荐准确性的影响",
        output_path=acc_png,
    )
    plot_metric(
        df,
        metric="overall_diversity",
        ylabel="总体多样性",
        title="风险系数 alpha 对总体多样性的影响",
        output_path=div_png,
    )

    print("\nSaved outputs:", flush=True)
    print(f"  {csv_path}", flush=True)
    print(f"  {xlsx_path}", flush=True)
    print(f"  {acc_png}", flush=True)
    print(f"  {div_png}", flush=True)


def plot_metric(df: pd.DataFrame, metric: str, ylabel: str, title: str, output_path: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(12, 4.6), dpi=180, sharex=True)
    for ax, dataset in zip(axes, ["OP", "Yelp"]):
        sub = df[df["dataset"] == dataset].sort_values("alpha")
        ax.plot(
            sub["alpha"],
            sub[metric],
            marker="o",
            linewidth=2.0,
            markersize=4.5,
            color="#1f77b4" if dataset == "OP" else "#b45f06",
        )
        ax.set_title(f"{dataset} 数据集")
        ax.set_xlabel("风险系数 alpha")
        ax.set_ylabel(ylabel)
        ax.set_xticks(ALPHAS)
        ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.35)

    fig.suptitle(title, fontsize=13)
    fig.tight_layout()
    fig.savefig(output_path, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    all_results: list[dict] = []
    for dataset in ["OP", "Yelp"]:
        print(f"\nLoading {dataset} scores", flush=True)
        scores = load_scores(dataset)
        print(f"{dataset} shape: {scores.shape}", flush=True)
        for alpha in ALPHAS:
            all_results.append(run_one_alpha(dataset, scores, alpha))
            pd.DataFrame(all_results).to_csv(
                OUT_DIR / "exact_alpha_sensitivity_results_partial.csv",
                index=False,
                encoding="utf-8-sig",
            )

    write_outputs(all_results)


if __name__ == "__main__":
    main()
