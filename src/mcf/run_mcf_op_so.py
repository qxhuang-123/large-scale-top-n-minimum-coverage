from __future__ import annotations

import argparse
import json
import math
import time
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from src.common.shared_experiment_inputs import ALPHA, D_PERCENTAGES, N, SEED, SHARED_INPUTS

try:
    from ortools.graph.python import min_cost_flow
except ImportError as exc:
    raise SystemExit(
        "This script requires OR-Tools. Install it with: python -m pip install ortools"
    ) from exc


ROOT = Path(r"C:\Users\24qxh\Documents\Codex\2026-07-11\new-chat-2")
OUT_DIR = ROOT / "outputs" / "network_flow_N10_D_percent_results"
OUT_XLSX = OUT_DIR / "network_flow_N10_D_percent_5datasets.xlsx"
OUT_JSON = OUT_DIR / "network_flow_N10_D_percent_5datasets.json"

SCORE_SCALE = 1_000_000


@dataclass(frozen=True)
class DatasetConfig:
    name: str
    scores_path: Path | None = None
    cand_path: Path | None = None
    source_path: Path | None = None
    source_kind: str | None = None
    cache_dir: Path | None = None
    sheet_name: str | None = None


DATASETS: dict[str, DatasetConfig] = {
    name: DatasetConfig(name=name, scores_path=spec.scores, cand_path=spec.candidates)
    for name, spec in SHARED_INPUTS.items()
}


def canonical_dataset_names(selected: list[str] | None) -> list[str]:
    if not selected:
        return list(DATASETS)
    aliases = {name.lower(): name for name in DATASETS}
    names = []
    for raw in selected:
        key = raw.lower()
        if key not in aliases:
            raise ValueError(f"Unknown dataset: {raw}. Choices: {', '.join(DATASETS)}")
        names.append(aliases[key])
    return names


def read_xlsx_matrix(path: Path, sheet_name: str | None = None) -> np.ndarray:
    print(f"Loading Excel matrix: {path}", flush=True)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    num_users = ws.max_row - 1
    num_items = ws.max_column - 1
    scores = np.empty((num_users, num_items), dtype=np.float32)
    t0 = time.time()
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        vals = row[1:]
        scores[r, :] = [np.nan if v is None or v == "" else float(v) for v in vals]
        if (r + 1) % 500 == 0:
            print(f"  read {r + 1:,}/{num_users:,} rows", flush=True)
    print(f"Loaded {scores.shape}, finite={np.isfinite(scores).sum():,}, {time.time() - t0:.1f}s", flush=True)
    return scores


def find_so_csv_parts(folder: Path) -> list[Path]:
    for prefix in ("SO_unobserved_R_ui_part", "SO_full_R_ui_part"):
        paths = sorted(folder.rglob(f"{prefix}*.csv"))
        by_name = {p.name: p for p in paths}
        required = [f"{prefix}{i}.csv" for i in range(1, 4)]
        missing = [name for name in required if name not in by_name]
        if not missing:
            return [by_name[name] for name in required]
    raise FileNotFoundError(
        f"Missing SO CSV parts under {folder}: expected either "
        "SO_unobserved_R_ui_part1/2/3.csv or SO_full_R_ui_part1/2/3.csv"
    )



def read_so_csv_parts(folder: Path) -> np.ndarray:
    existing = find_so_csv_parts(folder)

    frames = []
    for path in existing:
        print(f"Loading SO CSV part: {path}", flush=True)
        frames.append(pd.read_csv(path, index_col=0))
    df = pd.concat(frames, axis=1)
    scores = df.to_numpy(dtype=np.float32, copy=True)
    print(f"Loaded SO matrix {scores.shape}, finite={np.isfinite(scores).sum():,}", flush=True)
    return scores


def build_candidate_cache(scores: np.ndarray, cache_dir: Path, dataset_name: str) -> tuple[Path, Path]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    scores_path = cache_dir / f"{dataset_name.lower()}_scores_float32.npy"
    cand_path = cache_dir / f"{dataset_name.lower()}_cand_frac_0p4_seed{SEED}.npz"
    if scores_path.exists() and cand_path.exists():
        return scores_path, cand_path

    print(f"Building candidate cache for {dataset_name}: alpha={ALPHA}, N={N}", flush=True)
    np.save(scores_path, scores.astype(np.float32, copy=False))
    num_users, num_items = scores.shape
    user_indptr = np.zeros(num_users + 1, dtype=np.int64)
    user_items_parts: list[np.ndarray] = []
    item_users_parts: list[list[int]] = [[] for _ in range(num_items)]

    edge_count = 0
    t0 = time.time()
    for u in range(num_users):
        row = scores[u]
        valid = np.flatnonzero(np.isfinite(row))
        if valid.size:
            k = min(valid.size, max(N, int(math.ceil(ALPHA * valid.size))))
            if k < valid.size:
                local = np.argpartition(row[valid], -k)[-k:]
                cand = valid[local]
            else:
                cand = valid
            order = np.argsort(-row[cand], kind="mergesort")
            cand = cand[order].astype(np.int32, copy=False)
        else:
            cand = np.empty(0, dtype=np.int32)
        user_items_parts.append(cand)
        for item in cand:
            item_users_parts[int(item)].append(u)
        edge_count += int(cand.size)
        user_indptr[u + 1] = edge_count
        if (u + 1) % 500 == 0:
            print(f"  candidates for {u + 1:,}/{num_users:,} users", flush=True)

    user_indices = np.concatenate(user_items_parts).astype(np.int32, copy=False)
    item_indptr = np.zeros(num_items + 1, dtype=np.int64)
    item_users_flat_parts = []
    cursor = 0
    for item, users in enumerate(item_users_parts):
        arr = np.asarray(users, dtype=np.int32)
        item_users_flat_parts.append(arr)
        cursor += int(arr.size)
        item_indptr[item + 1] = cursor
    item_users = np.concatenate(item_users_flat_parts) if item_users_flat_parts else np.empty(0, dtype=np.int32)
    np.savez_compressed(
        cand_path,
        user_indptr=user_indptr,
        user_indices=user_indices,
        item_indptr=item_indptr,
        item_users=item_users,
    )
    print(f"Saved candidate cache edges={edge_count:,}, time={time.time() - t0:.1f}s", flush=True)
    return scores_path, cand_path


def ensure_cache(cfg: DatasetConfig) -> tuple[Path, Path]:
    if cfg.scores_path and cfg.cand_path and cfg.scores_path.exists() and cfg.cand_path.exists():
        return cfg.scores_path, cfg.cand_path
    raise FileNotFoundError(
        f"Shared cache missing for {cfg.name}. MCF must use the same immutable "
        "score/candidate files as TPCAR and BGCR; rebuild the shared cache first."
    )


def load_candidate_arrays(scores: np.ndarray, cand_path: Path):
    z = np.load(cand_path)
    user_indptr = z["user_indptr"].astype(np.int64, copy=False)
    item_key = "user_items" if "user_items" in z.files else "user_indices"
    user_items = z[item_key].astype(np.int32, copy=False)
    num_users = len(user_indptr) - 1
    num_items = scores.shape[1]
    user_quota = np.minimum(N, np.diff(user_indptr)).astype(np.int32, copy=False)
    item_counts = np.bincount(user_items, minlength=num_items)
    items_with_candidates = int(np.count_nonzero(item_counts))
    return user_indptr, user_items, user_quota, items_with_candidates


def canonical_top_n(
    scores: np.ndarray,
    user: int,
    items: np.ndarray,
    k: int,
) -> np.ndarray:
    """Select by score descending, breaking score ties by item ID ascending."""
    if k <= 0:
        return items[:0]
    if k >= len(items):
        return items

    user_scores = np.asarray(scores[user, items], dtype=np.float64)
    order = np.lexsort((items, -user_scores))
    return items[order[:k]]


def compute_naive(scores: np.ndarray, user_indptr: np.ndarray, user_items: np.ndarray, user_quota: np.ndarray):
    num_users, num_items = scores.shape
    counts = np.zeros(num_items, dtype=np.int32)
    total_score = 0.0
    total_slots = int(user_quota.sum())
    for u in range(num_users):
        s, e = int(user_indptr[u]), int(user_indptr[u + 1])
        k = int(user_quota[u])
        if k <= 0:
            continue
        candidates = user_items[s:e]
        selected = canonical_top_n(scores, u, candidates, k)
        counts[selected] += 1
        total_score += float(np.asarray(scores[u, selected], dtype=np.float64).sum())
    coverage = int(np.count_nonzero(counts))
    return {
        "objective": total_score,
        "accuracy": total_score / total_slots if total_slots else 0.0,
        "coverage": coverage,
    }



def solve_network_flow_for_d(
    scores: np.ndarray,
    user_indptr: np.ndarray,
    user_items: np.ndarray,
    user_quota: np.ndarray,
    d_target: int,
) -> dict:
    num_users, num_items = scores.shape
    total_slots = int(user_quota.sum())
    source = 0
    user0 = 1
    item0 = user0 + num_users
    cover_sink = item0 + num_items
    sink = cover_sink + 1
    num_nodes = sink + 1

    solver = min_cost_flow.SimpleMinCostFlow()

    active_users = np.flatnonzero(user_quota > 0).astype(np.int64, copy=False)
    solver.add_arcs_with_capacity_and_unit_cost(
        np.full(active_users.size, source, dtype=np.int64),
        user0 + active_users,
        user_quota[active_users].astype(np.int64, copy=False),
        np.zeros(active_users.size, dtype=np.int64),
    )

    edge_count = int(len(user_items))
    edge_starts = np.empty(edge_count, dtype=np.int64)
    cursor = 0
    for u in range(num_users):
        s, e = int(user_indptr[u]), int(user_indptr[u + 1])
        if e > s:
            edge_starts[s:e] = user0 + u
            cursor = e
    if cursor != edge_count:
        raise RuntimeError("Candidate edge cursor mismatch while building flow graph")

    edge_ends = item0 + user_items.astype(np.int64, copy=False)
    edge_costs = np.empty(edge_count, dtype=np.int64)
    for u in range(num_users):
        s, e = int(user_indptr[u]), int(user_indptr[u + 1])
        if e > s:
            items = user_items[s:e]
            edge_costs[s:e] = -np.rint(np.asarray(scores[u, items], dtype=np.float64) * SCORE_SCALE).astype(np.int64)
    edge_arc_ids = solver.add_arcs_with_capacity_and_unit_cost(
        edge_starts,
        edge_ends,
        np.ones(edge_count, dtype=np.int64),
        edge_costs,
    )

    item_nodes = item0 + np.arange(num_items, dtype=np.int64)
    # Exactly D units must pass through distinct first-copy item arcs. Remaining
    # selected copies go directly to the ordinary sink. This is the network-flow
    # form of z[i] <= sum_u x[u,i] and sum_i z[i] >= D.
    solver.add_arcs_with_capacity_and_unit_cost(
        item_nodes,
        np.full(num_items, cover_sink, dtype=np.int64),
        np.ones(num_items, dtype=np.int64),
        np.zeros(num_items, dtype=np.int64),
    )
    solver.add_arcs_with_capacity_and_unit_cost(
        item_nodes,
        np.full(num_items, sink, dtype=np.int64),
        np.full(num_items, max(0, total_slots), dtype=np.int64),
        np.zeros(num_items, dtype=np.int64),
    )

    supplies = [0] * (num_nodes + 1)
    supplies[source] = total_slots
    supplies[sink] = -(total_slots - d_target)
    supplies[cover_sink] = -d_target
    for node, supply in enumerate(supplies):
        solver.set_node_supply(node, int(supply))

    t0 = time.time()
    status = solver.solve()
    solve_time = time.time() - t0
    if status != solver.OPTIMAL:
        return {
            "status": int(status),
            "solve_time": solve_time,
            "objective": None,
            "accuracy": None,
            "coverage": None,
            "selected_edges": None,
        }

    objective_scaled = 0
    item_flow = np.zeros(num_items, dtype=np.int32)
    selected_edges = 0
    for arc in edge_arc_ids:
        flow = solver.flow(int(arc))
        if flow <= 0:
            continue
        end = solver.head(int(arc))
        cost = solver.unit_cost(int(arc))
        selected_edges += flow
        item_flow[end - item0] += flow
        objective_scaled += -cost * flow

    objective = objective_scaled / SCORE_SCALE
    return {
        "status": "OPTIMAL",
        "solve_time": solve_time,
        "objective": objective,
        "accuracy": objective / total_slots if total_slots else 0.0,
        "coverage": int(np.count_nonzero(item_flow)),
        "selected_edges": int(selected_edges),
    }


def estimate_arcs(num_users: int, num_items: int, candidate_edges: int) -> int:
    # source-user arcs + user-item arcs + item-cover-demand arcs + item-sink arcs
    return num_users + candidate_edges + 2 * num_items


def run_dataset(name: str, cfg: DatasetConfig, dry_run: bool = False, d_percentages: list[int] | None = None, skip_cache_build: bool = False) -> tuple[list[dict], dict, dict]:
    print(f"\n=== {name} ===", flush=True)
    pct_values = d_percentages or D_PERCENTAGES
    cache_ready = bool(cfg.scores_path and cfg.cand_path and cfg.scores_path.exists() and cfg.cand_path.exists())
    if (dry_run or skip_cache_build) and not cache_ready:
        if cfg.name == "SO" and cfg.source_path is not None:
            parts = find_so_csv_parts(cfg.source_path)
            meta = {
                "dataset": name,
                "N": N,
                "alpha": ALPHA,
                "source_parts": "; ".join(str(p) for p in parts),
                "source_bytes": sum(p.stat().st_size for p in parts),
                "engine": "ortools_simple_min_cost_flow",
                "dry_run_note": "SO source CSV parts are present; run without --dry-run to build cache and solve.",
            }
            status = "DRY_RUN_SOURCE_ONLY" if dry_run else "SKIPPED_CACHE_NOT_BUILT"
            rows = [{
                "数据集": name,
                "D比例": f"D-{pct}%",
                "D_target": None,
                "model_D_target": None,
                "迭代次数": iteration,
                "总体多样性": None,
                "目标函数值": None,
                "准确多样性": None,
                "时间s": None,
                "target_feasible": None,
                "status": status,
            } for iteration, pct in enumerate(pct_values)]
            summary = {"数据集": name, "N": N, "总时间s": 0.0, "total_recs": None}
            print(f"SO source parts present: {len(parts)} files, {meta['source_bytes']:,} bytes", flush=True)
            return rows, summary, meta
    scores_path, cand_path = ensure_cache(cfg)
    scores = np.load(scores_path, mmap_mode="r")
    user_indptr, user_items, user_quota, items_with_candidates = load_candidate_arrays(scores, cand_path)
    num_users, num_items = scores.shape
    total_slots = int(user_quota.sum())
    candidate_edges = int(len(user_items))
    arc_estimate = estimate_arcs(num_users, num_items, candidate_edges)
    start_all = time.time()
    naive = None if dry_run else compute_naive(scores, user_indptr, user_items, user_quota)
    naive_time = 0.0 if dry_run else time.time() - start_all

    meta = {
        "dataset": name,
        "N": N,
        "alpha": ALPHA,
        "num_users": int(num_users),
        "num_items": int(num_items),
        "candidate_edges": candidate_edges,
        "estimated_network_arcs": arc_estimate,
        "items_with_candidates": items_with_candidates,
        "total_slots": total_slots,
        "D_target_rounding": "ceil",
        "top_n_tie_break": "score_desc_item_id_asc",
        "scores_path": str(scores_path),
        "cand_path": str(cand_path),
        "engine": "ortools_simple_min_cost_flow",
    }
    print(
        f"shape={num_users:,}x{num_items:,}, slots={total_slots:,}, "
        f"candidate_edges={candidate_edges:,}, estimated_arcs={arc_estimate:,}",
        flush=True,
    )

    rows = []
    last_good = None
    for iteration, pct in enumerate(pct_values):
        # In the paper, coverage is imposed on the screened candidate item set I^alpha.
        raw_target = int(math.ceil(items_with_candidates * pct / 100.0))
        d_target = min(raw_target, items_with_candidates, total_slots)
        feasible_by_candidates = raw_target <= min(items_with_candidates, total_slots)
        if dry_run:
            row = {
                "数据集": name,
                "D比例": f"D-{pct}%",
                "D_target": raw_target,
                "D_target_base": "items_with_candidates",
                "D_target_rounding": "ceil",
                "top_n_tie_break": "score_desc_item_id_asc",
                "model_D_target": d_target,
                "迭代次数": iteration,
                "总体多样性": None,
                "目标函数值": None,
                "准确多样性": None,
                "时间s": None,
                "target_feasible": feasible_by_candidates,
                "status": "DRY_RUN",
            }
            rows.append(row)
            continue

        print(f"Solving {name} D-{pct}%: raw={raw_target:,}, model={d_target:,}", flush=True)
        if raw_target == 0:
            # The unconstrained problem decomposes by user. Returning the
            # canonical Top-N optimum prevents solver-dependent tie choices.
            result = {
                "status": "CANONICAL_TOP_N",
                "solve_time": naive_time,
                "objective": naive["objective"],
                "accuracy": naive["accuracy"],
                "coverage": naive["coverage"],
                "selected_edges": total_slots,
            }
        else:
            result = solve_network_flow_for_d(scores, user_indptr, user_items, user_quota, d_target)
        elapsed = time.time() - start_all
        target_feasible = bool(feasible_by_candidates and result["coverage"] is not None and result["coverage"] >= raw_target)
        row = {
            "数据集": name,
            "D比例": f"D-{pct}%",
            "D_target": raw_target,
            "D_target_base": "items_with_candidates",
            "D_target_rounding": "ceil",
            "top_n_tie_break": "score_desc_item_id_asc",
            "model_D_target": d_target,
            "迭代次数": iteration,
            "总体多样性": result["coverage"],
            "目标函数值": result["objective"],
            "准确多样性": result["accuracy"],
            "时间s": elapsed,
            "单次求解时间s": result["solve_time"],
            "target_feasible": target_feasible,
            "status": result["status"],
        }
        rows.append(row)
        if result["objective"] is not None:
            last_good = row
        print(
            f"  status={result['status']}, coverage={result['coverage']}, "
            f"objective={result['objective']}, elapsed={elapsed:.1f}s",
            flush=True,
        )

    summary = {
        "数据集": name,
        "N": N,
        "总时间s": time.time() - start_all,
        "目标函数值": None if last_good is None else last_good["目标函数值"],
        "准确多样性": None if last_good is None else last_good["准确多样性"],
        "总体多样性": None if last_good is None else last_good["总体多样性"],
        "Naive目标函数值": None if naive is None else naive["objective"],
        "Naive准确多样性": None if naive is None else naive["accuracy"],
        "Naive总体多样性": None if naive is None else naive["coverage"],
        "target_feasible": None if last_good is None else last_good["target_feasible"],
        "max_reached": None if last_good is None else last_good["总体多样性"],
        "total_recs": total_slots,
    }
    return rows, summary, meta


def write_outputs(d_rows: list[dict], summary_rows: list[dict], meta_rows: list[dict]):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(d_rows).to_excel(writer, sheet_name="D比例结果", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="总结果", index=False)
        pd.DataFrame(meta_rows).to_excel(writer, sheet_name="meta", index=False)

    wb = openpyxl.load_workbook(OUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 42)
    wb.save(OUT_XLSX)

    OUT_JSON.write_text(
        json.dumps(
            {"percent_results": d_rows, "summary": summary_rows, "meta": meta_rows},
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    print(f"Saved: {OUT_XLSX}", flush=True)
    print(f"Saved: {OUT_JSON}", flush=True)


def main():
    parser = argparse.ArgumentParser(description="Run classic min-cost-flow Top-N coverage solver for 5 datasets.")
    parser.add_argument("--datasets", nargs="*", help="Subset to run, e.g. --datasets OP Yelp")
    parser.add_argument("--dry-run", action="store_true", help="Only inspect cache/schema and write feasibility metadata.")
    parser.add_argument("--d-percentages", nargs="*", type=int, default=None, help="Override D percentages, e.g. --d-percentages 0 20")
    parser.add_argument("--skip-so-cache-build", action="store_true", help="Do not build SO cache from CSV when it is missing.")
    args = parser.parse_args()

    names = canonical_dataset_names(args.datasets)
    all_rows: list[dict] = []
    all_summary: list[dict] = []
    all_meta: list[dict] = []
    for name in names:
        rows, summary, meta = run_dataset(name, DATASETS[name], dry_run=args.dry_run, d_percentages=args.d_percentages, skip_cache_build=args.skip_so_cache_build and name == "SO")
        all_rows.extend(rows)
        all_summary.append(summary)
        all_meta.append(meta)
        write_outputs(all_rows, all_summary, all_meta)


if __name__ == "__main__":
    main()
