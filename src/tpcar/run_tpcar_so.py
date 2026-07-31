from __future__ import annotations

import argparse
import json
import sys
import time
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
from numpy.lib import format as npy_format
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, Side

try:
    import matplotlib.pyplot as plt
except Exception:
    plt = None

try:
    from numba import njit
except Exception:
    njit = None


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
CACHE_DIR = REPO_ROOT / "cache" / "so"
SCORES_PATH = CACHE_DIR / "SO_scores_float32.npy"
CANDIDATES_PATH = CACHE_DIR / "SO_cand_pos_top40_seed20260704_v2.npz"
DATA_PATH = REPO_ROOT / "data" / "raw" / "so"
OUTPUT_DIR = REPO_ROOT / "outputs" / "tpcar" / "so"

DATASET = "SO"
DEFAULT_N = 10
DEFAULT_PROGRESS_EVERY = 200
D_PCTS = [0, 20, 40]
CANDIDATE_FRACTION = 0.40
SEED = 20260704

try:
    import tpcar_core_fast as tpcar_core
except ImportError as exc:
    raise ImportError(
        "Cannot import tpcar_core_fast. Build it in src/tpcar with "
        "`python setup.py build_ext --inplace`."
    ) from exc


if njit is not None:

    @njit(cache=False)
    def _build_edge_maps_numba(user_indptr, user_items, item_indptr):
        num_users = len(user_indptr) - 1
        num_edges = len(user_items)
        edge_users = np.empty(num_edges, dtype=np.int32)
        item_edges = np.empty(num_edges, dtype=np.int32)
        cursor = item_indptr[:-1].copy()

        for user in range(num_users):
            start = user_indptr[user]
            end = user_indptr[user + 1]
            for edge in range(start, end):
                edge_users[edge] = user
                item = user_items[edge]
                pos = cursor[item]
                item_edges[pos] = edge
                cursor[item] = pos + 1
        return edge_users, item_edges


def _read_npy_header(stream) -> tuple[tuple[int, ...], np.dtype]:
    version = npy_format.read_magic(stream)
    shape, _, dtype = npy_format._read_array_header(stream, version)
    return shape, dtype


def inspect_inputs() -> dict:
    missing = [str(path) for path in (SCORES_PATH, CANDIDATES_PATH) if not path.exists()]
    if missing:
        raise FileNotFoundError("缺少 SO 精确算法缓存:\n" + "\n".join(missing))

    scores = np.load(SCORES_PATH, mmap_mode="r")
    headers: dict[str, dict] = {}
    with zipfile.ZipFile(CANDIDATES_PATH) as archive:
        members = set(archive.namelist())
        required = {
            "user_indptr.npy",
            "user_items.npy",
            "user_scores.npy",
            "item_indptr.npy",
        }
        absent = sorted(required - members)
        if absent:
            raise ValueError(f"候选缓存缺少数组: {absent}")
        for member in sorted(required):
            with archive.open(member) as stream:
                shape, dtype = _read_npy_header(stream)
            headers[member.removesuffix(".npy")] = {
                "shape": list(shape),
                "dtype": str(dtype),
            }

    num_users, num_items = map(int, scores.shape)
    edge_count = int(headers["user_items"]["shape"][0])
    if headers["user_indptr"]["shape"] != [num_users + 1]:
        raise ValueError("user_indptr 长度与 SO 用户数不一致")
    if headers["item_indptr"]["shape"] != [num_items + 1]:
        raise ValueError("item_indptr 长度与 SO 物品数不一致")
    if headers["user_scores"]["shape"] != [edge_count]:
        raise ValueError("user_scores 与 user_items 的长度不一致")
    if edge_count > np.iinfo(np.int32).max:
        raise OverflowError("候选边数量超过当前 C++ 精确算法的 int32 索引范围")

    return {
        "scores_shape": [num_users, num_items],
        "scores_dtype": str(scores.dtype),
        "candidate_edges": edge_count,
        "candidate_arrays": headers,
        "scores_path": str(SCORES_PATH),
        "candidates_path": str(CANDIDATES_PATH),
        "cpp_module": str(Path(tpcar_core.__file__).resolve()),
    }


def load_solver_arrays():
    if njit is None:
        raise RuntimeError("当前环境没有 numba，无法高效构造 SO 的 2.61 亿条候选边索引。")

    scores = np.load(SCORES_PATH, mmap_mode="r")
    candidates = np.load(CANDIDATES_PATH)
    try:
        user_indptr_64 = candidates["user_indptr"]
        item_indptr_64 = candidates["item_indptr"]
        if user_indptr_64[-1] > np.iinfo(np.int32).max:
            raise OverflowError("SO 候选边数量超过 int32 索引范围")

        user_indptr = user_indptr_64.astype(np.int32, copy=False)
        item_indptr = item_indptr_64.astype(np.int32, copy=False)
        user_items = candidates["user_items"].astype(np.int32, copy=False)

        print("converting cached user scores to float64", flush=True)
        user_scores = candidates["user_scores"].astype(np.float64, copy=False)
    finally:
        candidates.close()

    print("building compact C++ edge maps with numba", flush=True)
    started = time.time()
    edge_users, item_edges = _build_edge_maps_numba(user_indptr, user_items, item_indptr)
    build_seconds = time.time() - started
    print(f"edge maps built in {build_seconds:.3f}s", flush=True)

    return scores, (
        user_indptr,
        user_items,
        user_scores,
        edge_users,
        item_indptr,
        item_edges,
    ), build_seconds


def d_target_from_pct(d_base: int, pct: int) -> int:
    return int(round(d_base * pct / 100.0))


def choose_first_history_at_or_after(history: list[dict], target_d: int) -> dict:
    for row in history:
        if int(row["diversity"]) >= target_d:
            return row
    return history[-1]


def make_history_rows(result: dict, total_recs: int, n: int) -> list[dict]:
    rows = []
    for entry in result["history"]:
        objective = float(entry["objective"])
        rows.append(
            {
                "数据集": DATASET,
                "N": n,
                "迭代次数": int(entry["iteration"]),
                "总体多样性": int(entry["diversity"]),
                "目标函数值": round(objective, 6),
                "准确多样性": round(objective / total_recs, 6) if total_recs else 0.0,
                "时间s": round(float(entry["time_sec"]), 3),
            }
        )
    return rows


def make_percent_rows(
    result: dict,
    d_base: int,
    model_d_target: int,
    total_elapsed: float,
    n: int,
    d_pcts: list[int],
) -> list[dict]:
    history = list(result["history"])
    total_recs = int(result["total_recs"])
    rows = []
    for pct in d_pcts:
        target_d = d_target_from_pct(d_base, pct)
        entry = choose_first_history_at_or_after(history, target_d)
        objective = float(entry["objective"])
        achieved_d = int(entry["diversity"])
        feasible = achieved_d >= target_d
        rows.append(
            {
                "数据集": DATASET,
                "N": n,
                "D比例": f"D-{pct}%",
                "D_target": target_d,
                "D_target_base": "items_with_candidates",
                "model_D_target": model_d_target,
                "迭代次数": int(entry["iteration"]),
                "总体多样性": achieved_d,
                "目标函数值": round(objective, 6),
                "准确多样性": round(objective / total_recs, 6) if total_recs else 0.0,
                "时间s": round(float(entry["time_sec"]), 3),
                "target_feasible": feasible,
                "status": "TARGET_REACHED" if feasible else "MAX_REACHED",
            }
        )
    rows[-1]["时间s"] = round(total_elapsed, 3)
    return rows


def output_paths(n: int) -> tuple[Path, Path, Path, Path]:
    out_dir = OUTPUT_DIR
    stem = f"SO精确算法_N{n}_D_percent_history_results"
    return out_dir, out_dir / f"{stem}.json", out_dir / f"{stem}.xlsx", out_dir / f"SO精确算法_N{n}_iteration_curves.png"


def write_excel(path: Path, meta: dict, percent_rows: list[dict], history_rows: list[dict], summary_rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(history_rows).to_excel(writer, sheet_name="迭代明细", index=False)
        pd.DataFrame(percent_rows).to_excel(writer, sheet_name="D比例结果", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="总结果", index=False)
        pd.DataFrame([meta]).to_excel(writer, sheet_name="meta", index=False)

    workbook = load_workbook(path)
    thin = Side(style="thin", color="000000")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(name="宋体", size=11, bold=True)
        for column in sheet.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 24)

    history_sheet = workbook["迭代明细"]
    if history_sheet.max_row >= 2:
        categories = Reference(history_sheet, min_col=3, min_row=2, max_row=history_sheet.max_row)
        diversity_chart = LineChart()
        diversity_chart.title = "SO 总体多样性迭代曲线"
        diversity_chart.y_axis.title = "总体多样性"
        diversity_chart.x_axis.title = "迭代次数"
        diversity_chart.add_data(
            Reference(history_sheet, min_col=4, min_row=1, max_row=history_sheet.max_row),
            titles_from_data=True,
        )
        diversity_chart.set_categories(categories)
        diversity_chart.height = 7
        diversity_chart.width = 14
        history_sheet.add_chart(diversity_chart, "H2")

        objective_chart = LineChart()
        objective_chart.title = "SO 目标函数值迭代曲线"
        objective_chart.y_axis.title = "目标函数值"
        objective_chart.x_axis.title = "迭代次数"
        objective_chart.add_data(
            Reference(history_sheet, min_col=5, min_row=1, max_row=history_sheet.max_row),
            titles_from_data=True,
        )
        objective_chart.set_categories(categories)
        objective_chart.height = 7
        objective_chart.width = 14
        history_sheet.add_chart(objective_chart, "H18")
    workbook.save(path)


def write_figure(path: Path, history_rows: list[dict], n: int):
    if plt is None:
        print("matplotlib not installed; PNG skipped. Excel charts are still written.", flush=True)
        return
    if not history_rows:
        return
    frame = pd.DataFrame(history_rows).sort_values("迭代次数")
    figure, axes = plt.subplots(1, 2, figsize=(12, 4), dpi=160)
    axes[0].plot(frame["迭代次数"], frame["总体多样性"], linewidth=1.8)
    axes[0].set_title(f"SO N={n} 总体多样性")
    axes[0].set_xlabel("迭代次数")
    axes[0].set_ylabel("总体多样性 D")
    axes[0].grid(True, alpha=0.25)
    axes[1].plot(frame["迭代次数"], frame["目标函数值"], linewidth=1.8, color="#b45f06")
    axes[1].set_title(f"SO N={n} 目标函数值")
    axes[1].set_xlabel("迭代次数")
    axes[1].set_ylabel("目标函数值")
    axes[1].grid(True, alpha=0.25)
    figure.tight_layout()
    figure.savefig(path)
    plt.close(figure)


def parse_args():
    parser = argparse.ArgumentParser(description="SO 数据集 C++ 精确算法，输出 D 比例结果和完整迭代历史。")
    parser.add_argument("--n", type=int, default=DEFAULT_N, help="每名用户的推荐数量，默认 10")
    parser.add_argument("--progress-every", type=int, default=DEFAULT_PROGRESS_EVERY)
    parser.add_argument(
        "--d-percentages",
        nargs="+",
        type=int,
        default=D_PCTS,
        help="需要求解的 D/Ialpha 百分比，例如 --d-percentages 0 或 0 20 40 60",
    )
    parser.add_argument("--validate-only", action="store_true", help="只检查缓存、维度和 C++ 模块，不分配大型求解数组")
    return parser.parse_args()


def main():
    args = parse_args()
    if args.n <= 0:
        raise ValueError("--n 必须大于 0")
    if args.progress_every <= 0:
        raise ValueError("--progress-every 必须大于 0")
    if not args.d_percentages or any(pct < 0 or pct > 100 for pct in args.d_percentages):
        raise ValueError("--d-percentages 必须是 0 到 100 之间的整数")
    d_pcts = sorted(set(args.d_percentages))

    inspection = inspect_inputs()
    print(json.dumps(inspection, ensure_ascii=False, indent=2), flush=True)
    if args.validate_only:
        print("SO exact input validation passed.", flush=True)
        return

    scores, arrays, array_build_sec = load_solver_arrays()
    num_users, num_items = map(int, scores.shape)
    user_indptr, user_items, _, _, item_indptr, _ = arrays
    candidate_edges = int(len(user_items))
    items_with_candidates = int(np.count_nonzero(np.diff(item_indptr) > 0))
    d_base = items_with_candidates
    model_d_target = d_target_from_pct(d_base, max(d_pcts))
    avg_candidates = candidate_edges / num_users if num_users else 0.0
    print(
        f"candidate universe: full_items={num_items:,}, "
        f"items_with_candidates={items_with_candidates:,}, "
        f"edges={candidate_edges:,}, avg_per_user={avg_candidates:.2f}",
        flush=True,
    )

    meta = {
        "dataset": DATASET,
        "N": args.n,
        "num_users": num_users,
        "num_items": num_items,
        "candidate_edges": candidate_edges,
        "items_with_candidates": items_with_candidates,
        "D_percentages": d_pcts,
        "D_target_base": "items_with_candidates",
        "model_D_target": model_d_target,
        "candidate_fraction": CANDIDATE_FRACTION,
        "seed": SEED,
        "engine": "cpp_pybind11_exact_with_history",
        "data_path": str(DATA_PATH),
        "scores_path": str(SCORES_PATH),
        "candidates_path": str(CANDIDATES_PATH),
        "cpp_module": str(Path(tpcar_core.__file__).resolve()),
        "array_build_sec": round(array_build_sec, 3),
    }

    print(f"--- EXACT {DATASET} N={args.n}, D={model_d_target} ---", flush=True)
    started = time.time()
    raw = dict(
        tpcar_core.run_exact_csr(
            num_users,
            num_items,
            *arrays,
            args.n,
            model_d_target,
            args.progress_every,
        )
    )
    total_elapsed = time.time() - started

    history_rows = make_history_rows(raw, int(raw["total_recs"]), args.n)
    percent_rows = make_percent_rows(
        raw,
        d_base,
        model_d_target,
        total_elapsed,
        args.n,
        d_pcts,
    )
    feasible = bool(raw["target_feasible"])
    summary_rows = [
        {
            "数据集": DATASET,
            "N": args.n,
            "总时间s": round(total_elapsed, 3),
            "目标函数值": round(float(raw["tpcar_obj"]), 6),
            "准确多样性": round(float(raw["tpcar_pred"]), 6),
            "总体多样性": int(raw["tpcar_D"]),
            "Naive目标函数值": round(float(raw["naive_obj"]), 6),
            "Naive准确多样性": round(float(raw["naive_pred"]), 6),
            "Naive总体多样性": int(raw["naive_D"]),
            "迭代次数": int(raw["augmentations"]),
            "交换次数": int(raw["swaps"]),
            "D_target_base": "items_with_candidates",
            "model_D_target": model_d_target,
            "target_feasible": feasible,
            "status": "TARGET_REACHED" if feasible else "MAX_REACHED",
            "max_reached": int(raw["max_reached"]),
            "total_recs": int(raw["total_recs"]),
        }
    ]

    out_dir, out_json, out_xlsx, out_figure = output_paths(args.n)
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": meta,
        "D_percent_results": percent_rows,
        "summary": summary_rows,
        "history": history_rows,
    }
    out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(out_xlsx, meta, percent_rows, history_rows, summary_rows)
    write_figure(out_figure, history_rows, args.n)

    print("D percent results:", flush=True)
    for row in percent_rows:
        print(row, flush=True)
    print("summary:", summary_rows[0], flush=True)
    print(out_json, flush=True)
    print(out_xlsx, flush=True)
    if out_figure.exists():
        print(out_figure, flush=True)


if __name__ == "__main__":
    main()
